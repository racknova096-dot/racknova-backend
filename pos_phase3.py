from __future__ import annotations
from uuid import UUID
from fastapi import Header
from multiempresa_tenant import bind_empresa as _rn_bind_empresa
import multiempresa_tenant as rn_tenant

import io
import json
import os
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from typing import Any, Callable, Dict, List, Optional, Tuple
from uuid import uuid4

from fastapi import Depends, HTTPException, Query, Response, status
from pydantic import BaseModel, Field as PydanticField
from sqlalchemy import or_
from sqlmodel import Field, Session, SQLModel, select
from sqlalchemy import text as sa_text

# La Fase 3 es aditiva: reutiliza las tablas y modelos estables de la Fase 2.
try:
    from pos_module import (
        POSCaja,
        POSDevolucion,
        POSDevolucionDetalle,
        POSMovimientoEfectivo,
        POSSesionCaja,
        POSVentaControl,
        POSVentaLote,
        POSVentaMovimiento,
        VentaPOS,
        VentaPOSDetalle,
        VentaPOSPago,
    )
except ImportError as exc:  # pragma: no cover - mensaje útil durante despliegue
    raise RuntimeError(
        "RackNova POS Fase 3 requiere que pos_module.py de la Fase 2 esté instalado."
    ) from exc


# ==========================================================
# MODELOS — FASE 3
# ==========================================================


class POSCliente(SQLModel, table=True):
    __tablename__ = "pos_cliente"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_cliente: Optional[int] = Field(default=None, primary_key=True)
    nombre: str = Field(index=True)
    telefono: Optional[str] = Field(default=None, index=True)
    email: Optional[str] = Field(default=None, index=True)
    rfc: Optional[str] = Field(default=None, index=True)
    direccion: Optional[str] = None
    limite_credito: float = 0
    dias_credito: int = 0
    notas: Optional[str] = None
    activo: bool = Field(default=True, index=True)
    fecha_creacion: datetime
    fecha_actualizacion: datetime


class POSProductoConfiguracion(SQLModel, table=True):
    __tablename__ = "pos_producto_configuracion"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_configuracion: Optional[int] = Field(default=None, primary_key=True)
    sku: str = Field(index=True)
    unidad_venta: str = "pieza"
    permite_fraccion: bool = False
    # Unidades del inventario que equivalen a 1 unidad de venta.
    # Ejemplo: inventario en gramos y venta en kg => 1000.
    factor_inventario: float = 1
    precio_normal: Optional[float] = None
    precio_mayoreo: Optional[float] = None
    cantidad_mayoreo: float = 0
    precio_minimo: Optional[float] = None
    activo: bool = True
    fecha_actualizacion: datetime
    actualizado_por: str


class POSPrecioCliente(SQLModel, table=True):
    __tablename__ = "pos_precio_cliente"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_precio: Optional[int] = Field(default=None, primary_key=True)
    id_cliente: int = Field(index=True)
    sku: str = Field(index=True)
    precio: float
    activo: bool = True
    fecha_actualizacion: datetime
    actualizado_por: str


class POSPromocion(SQLModel, table=True):
    __tablename__ = "pos_promocion"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_promocion: Optional[int] = Field(default=None, primary_key=True)
    nombre: str
    tipo: str = Field(index=True)  # PORCENTAJE / PRECIO_FIJO / NXM
    sku: Optional[str] = Field(default=None, index=True)  # NULL = todos
    porcentaje: float = 0
    precio_fijo: float = 0
    cantidad_minima: float = 0
    compra_cantidad: float = 0
    paga_cantidad: float = 0
    fecha_inicio: Optional[datetime] = None
    fecha_fin: Optional[datetime] = None
    prioridad: int = 0
    activa: bool = Field(default=True, index=True)
    fecha_creacion: datetime
    creada_por: str


class POSVentaExtra(SQLModel, table=True):
    __tablename__ = "pos_venta_extra"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_extra: Optional[int] = Field(default=None, primary_key=True)
    id_venta: int = Field(index=True)
    id_cliente: Optional[int] = Field(default=None, index=True)
    cliente_nombre: Optional[str] = None
    tipo_venta: str = Field(default="CONTADO", index=True)
    saldo_pendiente: float = 0
    fecha_vencimiento: Optional[date] = None
    descuento_promociones: float = 0
    promociones_json: str = "[]"
    version_pos: str = "3"
    fecha_creacion: datetime


class POSVentaDetalleExtra(SQLModel, table=True):
    __tablename__ = "pos_venta_detalle_extra"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_extra_detalle: Optional[int] = Field(default=None, primary_key=True)
    id_detalle: int = Field(index=True)
    id_venta: int = Field(index=True)
    unidad_venta: str = "pieza"
    cantidad_venta: float = 0
    cantidad_inventario: int = 0
    factor_inventario: float = 1
    descuento_automatico: float = 0
    promocion_nombre: Optional[str] = None
    precio_origen: str = "LISTA"


class POSCredito(SQLModel, table=True):
    __tablename__ = "pos_credito"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_credito: Optional[int] = Field(default=None, primary_key=True)
    id_venta: int = Field(index=True)
    id_cliente: int = Field(index=True)
    total_credito: float
    saldo: float
    fecha_vencimiento: date
    estado: str = Field(default="PENDIENTE", index=True)
    usuario_autorizo: str
    fecha_creacion: datetime
    fecha_actualizacion: datetime


class POSAbono(SQLModel, table=True):
    __tablename__ = "pos_abono"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_abono: Optional[int] = Field(default=None, primary_key=True)
    folio: str = Field(index=True)
    id_credito: int = Field(index=True)
    id_cliente: int = Field(index=True)
    id_sesion: Optional[int] = Field(default=None, index=True)
    metodo: str = Field(index=True)
    monto: float
    referencia: Optional[str] = None
    usuario: str
    fecha: datetime


class POSDevolucionExtra(SQLModel, table=True):
    __tablename__ = "pos_devolucion_extra"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_extra: Optional[int] = Field(default=None, primary_key=True)
    id_devolucion: int = Field(index=True)
    ajuste_credito: float = 0
    reembolso_real: float = 0


class POSReporteDiario(SQLModel, table=True):
    __tablename__ = "pos_reporte_diario"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_reporte: Optional[int] = Field(default=None, primary_key=True)
    fecha_reporte: date = Field(index=True)
    cerrado: bool = False
    datos_json: str
    generado_por: str
    fecha_generacion: datetime


class POSAuditoria(SQLModel, table=True):
    __tablename__ = "pos_auditoria"
    # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST_MODEL
    empresa_id: UUID | None = Field(default=None, nullable=False, index=True)

    id_auditoria: Optional[int] = Field(default=None, primary_key=True)
    accion: str = Field(index=True)
    entidad: str = Field(index=True)
    entidad_id: Optional[str] = Field(default=None, index=True)
    detalles_json: str = "{}"
    usuario: str
    fecha: datetime


# ==========================================================
# REQUESTS
# ==========================================================


class ClienteRequest(BaseModel):
    nombre: str = PydanticField(min_length=2, max_length=150)
    telefono: Optional[str] = PydanticField(default=None, max_length=40)
    email: Optional[str] = PydanticField(default=None, max_length=150)
    rfc: Optional[str] = PydanticField(default=None, max_length=20)
    direccion: Optional[str] = PydanticField(default=None, max_length=300)
    limite_credito: float = PydanticField(default=0, ge=0)
    dias_credito: int = PydanticField(default=0, ge=0, le=3650)
    notas: Optional[str] = PydanticField(default=None, max_length=1000)
    activo: bool = True


class ProductoConfigRequest(BaseModel):
    sku: str = PydanticField(min_length=1, max_length=100)
    unidad_venta: str = PydanticField(default="pieza", min_length=1, max_length=30)
    permite_fraccion: bool = False
    factor_inventario: float = PydanticField(default=1, gt=0, le=1_000_000)
    precio_normal: Optional[float] = PydanticField(default=None, ge=0)
    precio_mayoreo: Optional[float] = PydanticField(default=None, ge=0)
    cantidad_mayoreo: float = PydanticField(default=0, ge=0)
    precio_minimo: Optional[float] = PydanticField(default=None, ge=0)
    activo: bool = True


class ProductoUnidadRequest(BaseModel):
    unidad_venta: str = PydanticField(min_length=1, max_length=30)


class PrecioClienteRequest(BaseModel):
    id_cliente: int = PydanticField(gt=0)
    sku: str = PydanticField(min_length=1, max_length=100)
    precio: float = PydanticField(gt=0)
    activo: bool = True


class PromocionRequest(BaseModel):
    nombre: str = PydanticField(min_length=2, max_length=150)
    tipo: str = PydanticField(min_length=2, max_length=30)
    sku: Optional[str] = PydanticField(default=None, max_length=100)
    porcentaje: float = PydanticField(default=0, ge=0, le=100)
    precio_fijo: float = PydanticField(default=0, ge=0)
    cantidad_minima: float = PydanticField(default=0, ge=0)
    compra_cantidad: float = PydanticField(default=0, ge=0)
    paga_cantidad: float = PydanticField(default=0, ge=0)
    fecha_inicio: Optional[datetime] = None
    fecha_fin: Optional[datetime] = None
    prioridad: int = PydanticField(default=0, ge=-1000, le=1000)
    activa: bool = True


class VentaItemV3Request(BaseModel):
    sku: str = PydanticField(min_length=1)
    cantidad: float = PydanticField(gt=0)
    descuento_porcentaje: float = PydanticField(default=0, ge=0, le=100)


class PagoV3Request(BaseModel):
    metodo: str = PydanticField(min_length=1)
    monto: float = PydanticField(gt=0)
    referencia: Optional[str] = PydanticField(default=None, max_length=150)


class VentaV3Request(BaseModel):
    operacion_id: str = PydanticField(min_length=8, max_length=100)
    items: List[VentaItemV3Request] = PydanticField(min_length=1)
    pagos: List[PagoV3Request] = PydanticField(default_factory=list)
    efectivo_recibido: Optional[float] = PydanticField(default=None, ge=0)
    id_cliente: Optional[int] = PydanticField(default=None, gt=0)
    tipo_venta: str = "CONTADO"  # CONTADO / CREDITO / PARCIAL
    fecha_vencimiento: Optional[date] = None


class AbonoRequest(BaseModel):
    monto: float = PydanticField(gt=0)
    metodo: str = PydanticField(min_length=1)
    referencia: Optional[str] = PydanticField(default=None, max_length=150)


class CerrarCajaV3Request(BaseModel):
    efectivo_contado: float = PydanticField(ge=0)
    observaciones: Optional[str] = PydanticField(default=None, max_length=500)


class CancelarV3Request(BaseModel):
    motivo: str = PydanticField(min_length=5, max_length=500)


class DevolucionItemV3Request(BaseModel):
    id_detalle: int = PydanticField(gt=0)
    cantidad: float = PydanticField(gt=0)


class DevolucionV3Request(BaseModel):
    items: List[DevolucionItemV3Request] = PydanticField(min_length=1)
    motivo: str = PydanticField(min_length=5, max_length=500)
    metodo_reembolso: str = PydanticField(default="efectivo")


# ==========================================================
# UTILIDADES
# ==========================================================



_PRODUCT_UNITS: Dict[str, Dict[str, Any]] = {
    "pieza": {
        "unidad_venta": "pieza",
        "etiqueta": "Pieza",
        "simbolo": "pza",
        "permite_fraccion": False,
        "factor_inventario": 1.0,
        "unidad_interna": "pieza",
    },
    "kg": {
        "unidad_venta": "kg",
        "etiqueta": "Kilogramo",
        "simbolo": "kg",
        "permite_fraccion": True,
        "factor_inventario": 1000.0,
        "unidad_interna": "gramo",
    },
    "litro": {
        "unidad_venta": "litro",
        "etiqueta": "Litro",
        "simbolo": "L",
        "permite_fraccion": True,
        "factor_inventario": 1000.0,
        "unidad_interna": "mililitro",
    },
}

_PRODUCT_UNIT_ALIASES = {
    "pieza": "pieza",
    "piezas": "pieza",
    "pza": "pieza",
    "pzas": "pieza",
    "unidad": "pieza",
    "unidades": "pieza",
    "kg": "kg",
    "kilo": "kg",
    "kilos": "kg",
    "kilogramo": "kg",
    "kilogramos": "kg",
    "l": "litro",
    "lt": "litro",
    "lts": "litro",
    "litro": "litro",
    "litros": "litro",
}


def _product_unit(value: Any) -> Tuple[str, Dict[str, Any]]:
    clean = str(value or "pieza").strip().lower()
    key = _PRODUCT_UNIT_ALIASES.get(clean)
    if not key:
        raise HTTPException(
            status_code=400,
            detail="Unidad inválida. Usa pieza, kg o litro.",
        )
    return key, _PRODUCT_UNITS[key]


def _serialize_product_unit(
    sku: str,
    row: Optional[POSProductoConfiguracion],
    *,
    product_exists: bool,
) -> Dict[str, Any]:
    key, definition = _product_unit(row.unidad_venta if row else "pieza")
    factor = float(row.factor_inventario if row else definition["factor_inventario"])
    return {
        "sku": sku,
        "producto_existe": product_exists,
        "unidad_venta": key,
        "etiqueta": definition["etiqueta"],
        "simbolo": definition["simbolo"],
        "unidad_interna": definition["unidad_interna"],
        "permite_fraccion": bool(
            row.permite_fraccion if row else definition["permite_fraccion"]
        ),
        "factor_inventario": factor,
        "activo": bool(row.activo) if row else True,
    }


def _dump(model: Any) -> Dict[str, Any]:
    if hasattr(model, "model_dump"):
        return dict(model.model_dump())
    return dict(model.dict())


def _money(value: Any) -> float:
    return round(float(value or 0), 2)


def _qty(value: Any) -> float:
    return round(float(value or 0), 6)


def _name(current_user: Any) -> str:
    return str(
        getattr(current_user, "nombre", None)
        or getattr(current_user, "usuario", None)
        or "Usuario"
    )


def _key(current_user: Any) -> str:
    return str(
        getattr(current_user, "usuario", None)
        or getattr(current_user, "nombre", None)
        or "Usuario"
    )


def _role(current_user: Any) -> str:
    return str(getattr(current_user, "_racknova_empresa_role", None) or getattr(current_user, "rol", "operator") or "operator").lower()


def _folio(prefix: str, now: datetime) -> str:
    return f"{prefix}-{now:%Y%m%d-%H%M%S}-{uuid4().hex[:6].upper()}"


def _audit(
    session: Session,
    *,
    action: str,
    entity: str,
    entity_id: Optional[Any],
    details: Dict[str, Any],
    user: str,
    now: datetime,
) -> None:
    session.add(
        POSAuditoria(
            accion=action,
            entidad=entity,
            entidad_id=None if entity_id is None else str(entity_id),
            detalles_json=json.dumps(details, ensure_ascii=False, default=str),
            usuario=user,
            fecha=now,
        )
    )


# RACKNOVA_POS_QUERY_OPTIMIZATIONS
def _open_session(session: Session, current_user: Any) -> Optional[POSSesionCaja]:
    keys = {
        value
        for value in (_key(current_user), _name(current_user))
        if value
    }

    if not keys:
        return None

    return session.exec(
        select(POSSesionCaja)
        .where(
            (POSSesionCaja.estado == "ABIERTA")
            & (POSSesionCaja.usuario.in_(keys))
        )
        .order_by(POSSesionCaja.fecha_apertura.desc())
    ).first()


def _require_open_session(session: Session, current_user: Any) -> POSSesionCaja:
    row = _open_session(session, current_user)
    if not row:
        raise HTTPException(
            status_code=400,
            detail="Debes abrir una caja antes de realizar esta operación.",
        )
    return row


def _client(session: Session, client_id: Optional[int]) -> Optional[POSCliente]:
    if client_id is None:
        return None
    row = session.get(POSCliente, client_id)
    if not row or not row.activo:
        raise HTTPException(status_code=404, detail="Cliente no encontrado o inactivo.")
    return row


def _product_config(
    session: Session,
    sku: str,
    *,
    create: bool = False,
    now: Optional[datetime] = None,
    user: str = "Sistema",
) -> Optional[POSProductoConfiguracion]:
    row = session.exec(
        select(POSProductoConfiguracion).where(POSProductoConfiguracion.sku == sku)
    ).first()
    if row or not create:
        return row
    row = POSProductoConfiguracion(
        sku=sku,
        unidad_venta="pieza",
        permite_fraccion=False,
        factor_inventario=1,
        activo=True,
        fecha_actualizacion=now or datetime.now(),
        actualizado_por=user,
    )
    session.add(row)
    session.flush()
    return row


def _active_promotions(
    session: Session,
    sku: str,
    now: datetime,
) -> List[POSPromocion]:
    cache_key = "_racknova_promociones_vigentes"
    rows = session.info.get(cache_key)

    if rows is None:
        rows = session.exec(
            select(POSPromocion)
            .where(
                (POSPromocion.activa == True)  # noqa: E712
                & (
                    (POSPromocion.fecha_inicio == None)  # noqa: E711
                    | (POSPromocion.fecha_inicio <= now)
                )
                & (
                    (POSPromocion.fecha_fin == None)  # noqa: E711
                    | (POSPromocion.fecha_fin >= now)
                )
            )
            .order_by(
                POSPromocion.prioridad.desc(),
                POSPromocion.id_promocion.asc(),
            )
        ).all()
        session.info[cache_key] = rows

    return [row for row in rows if not row.sku or row.sku == sku]


def _customer_price(
    session: Session,
    client_id: Optional[int],
    sku: str,
) -> Optional[POSPrecioCliente]:
    if client_id is None:
        return None
    return session.exec(
        select(POSPrecioCliente).where(
            (POSPrecioCliente.id_cliente == client_id)
            & (POSPrecioCliente.sku == sku)
            & (POSPrecioCliente.activo == True)  # noqa: E712
        )
    ).first()


def _calculate_price(
    session: Session,
    *,
    product: Any,
    config: Optional[POSProductoConfiguracion],
    client_id: Optional[int],
    quantity: float,
    manual_discount: float,
    now: datetime,
    is_admin: bool,
) -> Dict[str, Any]:
    factor = _qty(config.factor_inventario if config else 1) or 1
    base = _money(
        config.precio_normal
        if config and config.precio_normal is not None
        else product.precio_venta_sugerido
    )
    origin = "PRECIO_NORMAL" if config and config.precio_normal is not None else "LISTA"

    special = _customer_price(session, client_id, product.sku)
    if special:
        base = _money(special.precio)
        origin = "CLIENTE"
    elif (
        config
        and config.precio_mayoreo is not None
        and config.cantidad_mayoreo > 0
        and quantity >= config.cantidad_mayoreo
    ):
        base = _money(config.precio_mayoreo)
        origin = "MAYOREO"

    if base <= 0:
        raise HTTPException(
            status_code=400,
            detail=f"{product.nombre} no tiene precio de venta configurado.",
        )

    line_list = _money(base * quantity)
    automatic_total = line_list
    applied_name: Optional[str] = None
    applied_id: Optional[int] = None

    for promo in _active_promotions(session, product.sku, now):
        if quantity < float(promo.cantidad_minima or 0):
            continue
        candidate = line_list
        promo_type = promo.tipo.strip().upper()
        if promo_type == "PORCENTAJE" and promo.porcentaje > 0:
            candidate = _money(line_list * (1 - promo.porcentaje / 100))
        elif promo_type == "PRECIO_FIJO" and promo.precio_fijo > 0:
            candidate = _money(promo.precio_fijo * quantity)
        elif (
            promo_type == "NXM"
            and promo.compra_cantidad > 0
            and promo.paga_cantidad >= 0
            and promo.paga_cantidad < promo.compra_cantidad
        ):
            groups = int(quantity // promo.compra_cantidad)
            remainder = quantity - groups * promo.compra_cantidad
            payable = groups * promo.paga_cantidad + remainder
            candidate = _money(base * payable)
        else:
            continue

        if candidate < automatic_total:
            automatic_total = candidate
            applied_name = promo.nombre
            applied_id = promo.id_promocion
            origin = "PROMOCION"

    automatic_discount = _money(line_list - automatic_total)
    manual_amount = _money(automatic_total * (manual_discount / 100))
    final_total = _money(automatic_total - manual_amount)
    final_unit = _money(final_total / quantity)

    min_price = _money(config.precio_minimo) if config and config.precio_minimo else 0
    if not is_admin and min_price > 0 and final_unit < min_price:
        raise HTTPException(
            status_code=403,
            detail=(
                f"El precio final de {product.nombre} (${final_unit:.2f}) es menor "
                f"al mínimo autorizado (${min_price:.2f})."
            ),
        )

    cost_sale_unit = _money(float(product.costo_proveedor or 0) * factor)
    cost_total = _money(cost_sale_unit * quantity)

    return {
        "factor": factor,
        "base_unit": base,
        "line_list": line_list,
        "automatic_total": automatic_total,
        "automatic_discount": automatic_discount,
        "manual_discount_amount": manual_amount,
        "final_total": final_total,
        "final_unit": final_unit,
        "cost_unit": cost_sale_unit,
        "cost_total": cost_total,
        "profit": _money(final_total - cost_total),
        "promotion_name": applied_name,
        "promotion_id": applied_id,
        "origin": origin,
    }


def _detail_extra(session: Session, detail_id: int) -> Optional[POSVentaDetalleExtra]:
    return session.exec(
        select(POSVentaDetalleExtra).where(
            POSVentaDetalleExtra.id_detalle == detail_id
        )
    ).first()


def _returned_inventory(session: Session, detail_id: int) -> int:
    rows = session.exec(
        select(POSDevolucionDetalle).where(
            POSDevolucionDetalle.id_detalle_venta == detail_id
        )
    ).all()
    # En V3 guardamos POSDevolucionDetalle.cantidad en unidades de inventario.
    return int(sum(int(row.cantidad or 0) for row in rows))


def _sale_extra(session: Session, sale_id: int) -> Optional[POSVentaExtra]:
    return session.exec(
        select(POSVentaExtra).where(POSVentaExtra.id_venta == sale_id)
    ).first()


def _sale_control(session: Session, sale_id: int) -> Optional[POSVentaControl]:
    return session.exec(
        select(POSVentaControl).where(POSVentaControl.id_venta == sale_id)
    ).first()


def _credit_by_sale(session: Session, sale_id: int) -> Optional[POSCredito]:
    return session.exec(
        select(POSCredito).where(POSCredito.id_venta == sale_id)
    ).first()


def _serialize_clients_bulk(
    rows: List[POSCliente],
    session: Session,
) -> List[Dict[str, Any]]:
    client_ids = [
        int(row.id_cliente)
        for row in rows
        if row.id_cliente is not None
    ]

    balances: Dict[int, float] = defaultdict(float)
    overdue: Dict[int, float] = defaultdict(float)

    if client_ids:
        credits = session.exec(
            select(POSCredito).where(
                (POSCredito.id_cliente.in_(client_ids))
                & (POSCredito.estado != "CANCELADO")
                & (POSCredito.estado != "PAGADO")
            )
        ).all()

        today = date.today()
        for credit in credits:
            balances[credit.id_cliente] += float(credit.saldo or 0)
            if credit.fecha_vencimiento < today:
                overdue[credit.id_cliente] += float(credit.saldo or 0)

    result: List[Dict[str, Any]] = []
    for row in rows:
        client_id = int(row.id_cliente or 0)
        balance = _money(balances.get(client_id, 0))
        expired = _money(overdue.get(client_id, 0))

        result.append(
            {
                "id_cliente": row.id_cliente,
                "nombre": row.nombre,
                "telefono": row.telefono,
                "email": row.email,
                "rfc": row.rfc,
                "direccion": row.direccion,
                "limite_credito": _money(row.limite_credito),
                "dias_credito": row.dias_credito,
                "notas": row.notas,
                "activo": row.activo,
                "saldo": balance,
                "vencido": expired,
                "credito_disponible": _money(
                    max(row.limite_credito - balance, 0)
                ),
                "fecha_creacion": row.fecha_creacion,
                "fecha_actualizacion": row.fecha_actualizacion,
            }
        )

    return result


def _serialize_client(row: POSCliente, session: Optional[Session] = None) -> Dict[str, Any]:
    balance = 0.0
    overdue = 0.0
    if session and row.id_cliente is not None:
        credits = session.exec(
            select(POSCredito).where(POSCredito.id_cliente == row.id_cliente)
        ).all()
        balance = _money(sum(c.saldo for c in credits if c.estado not in {"CANCELADO", "PAGADO"}))
        today = date.today()
        overdue = _money(
            sum(
                c.saldo
                for c in credits
                if c.estado not in {"CANCELADO", "PAGADO"}
                and c.fecha_vencimiento < today
            )
        )
    return {
        "id_cliente": row.id_cliente,
        "nombre": row.nombre,
        "telefono": row.telefono,
        "email": row.email,
        "rfc": row.rfc,
        "direccion": row.direccion,
        "limite_credito": _money(row.limite_credito),
        "dias_credito": row.dias_credito,
        "notas": row.notas,
        "activo": row.activo,
        "saldo": balance,
        "vencido": overdue,
        "credito_disponible": _money(max(row.limite_credito - balance, 0)),
        "fecha_creacion": row.fecha_creacion,
        "fecha_actualizacion": row.fecha_actualizacion,
    }


def _serialize_sale(session: Session, sale: VentaPOS, detail: bool = True) -> Dict[str, Any]:
    extra = _sale_extra(session, int(sale.id_venta or 0))
    control = _sale_control(session, int(sale.id_venta or 0))
    data: Dict[str, Any] = {
        "id_venta": sale.id_venta,
        "folio": sale.folio,
        "usuario": sale.usuario,
        "subtotal": _money(sale.subtotal),
        "descuento_total": _money(sale.descuento_total),
        "total": _money(sale.total),
        "costo_total": _money(sale.costo_total),
        "ganancia": _money(sale.ganancia),
        "efectivo_recibido": _money(sale.efectivo_recibido),
        "cambio": _money(sale.cambio),
        "estado": sale.estado,
        "fecha": sale.fecha,
        "id_cliente": extra.id_cliente if extra else None,
        "cliente_nombre": extra.cliente_nombre if extra else None,
        "tipo_venta": extra.tipo_venta if extra else "CONTADO",
        "saldo_pendiente": _money(extra.saldo_pendiente if extra else 0),
        "fecha_vencimiento": extra.fecha_vencimiento if extra else None,
        "descuento_promociones": _money(extra.descuento_promociones if extra else 0),
        "id_sesion": control.id_sesion if control else None,
        "motivo_anulacion": control.motivo_anulacion if control else None,
        "fecha_anulacion": control.fecha_anulacion if control else None,
    }
    if not detail or sale.id_venta is None:
        return data

    items = session.exec(
        select(VentaPOSDetalle).where(VentaPOSDetalle.id_venta == sale.id_venta)
    ).all()
    result_items: List[Dict[str, Any]] = []
    for item in items:
        extra_item = _detail_extra(session, int(item.id_detalle or 0))
        quantity_inventory = (
            int(extra_item.cantidad_inventario)
            if extra_item
            else int(item.cantidad or 0)
        )
        returned_inventory = _returned_inventory(session, int(item.id_detalle or 0))
        factor = float(extra_item.factor_inventario if extra_item else 1) or 1
        qty_sale = (
            _qty(extra_item.cantidad_venta)
            if extra_item
            else _qty(item.cantidad)
        )
        qty_returned = _qty(returned_inventory / factor)
        result_items.append(
            {
                "id_detalle": item.id_detalle,
                "id_producto": item.id_producto,
                "sku": item.sku,
                "codigo_barras": item.codigo_barras,
                "nombre": item.nombre,
                "cantidad": qty_sale,
                "cantidad_devuelta": qty_returned,
                "cantidad_inventario": quantity_inventory,
                "unidad_venta": extra_item.unidad_venta if extra_item else "pieza",
                "factor_inventario": factor,
                "precio_lista": _money(item.precio_lista),
                "descuento_porcentaje": _money(item.descuento_porcentaje),
                "descuento_automatico": _money(
                    extra_item.descuento_automatico if extra_item else 0
                ),
                "promocion_nombre": extra_item.promocion_nombre if extra_item else None,
                "precio_origen": extra_item.precio_origen if extra_item else "LISTA",
                "precio_unitario_final": _money(item.precio_unitario_final),
                "subtotal": _money(item.subtotal),
                "costo_unitario": _money(item.costo_unitario),
                "costo_total": _money(item.costo_total),
                "ganancia": _money(item.ganancia),
            }
        )
    payments = session.exec(
        select(VentaPOSPago).where(VentaPOSPago.id_venta == sale.id_venta)
    ).all()
    data["items"] = result_items
    data["pagos"] = [
        {
            "id_pago": row.id_pago,
            "metodo": row.metodo,
            "monto": _money(row.monto),
            "referencia": row.referencia,
        }
        for row in payments
    ]
    return data


def _restore_lots(
    session: Session,
    ProductoLote: Any,
    sale_id: int,
    detail_id: int,
    amount_inventory: int,
) -> None:
    remaining = int(amount_inventory)
    links = session.exec(
        select(POSVentaLote)
        .where(
            (POSVentaLote.id_venta == sale_id)
            & (POSVentaLote.id_detalle == detail_id)
        )
        .order_by(POSVentaLote.id_registro.desc())
    ).all()
    for link in links:
        if remaining <= 0:
            break
        available = max(int(link.cantidad or 0) - int(link.cantidad_restaurada or 0), 0)
        restore = min(available, remaining)
        if restore <= 0:
            continue
        if link.id_lote is not None:
            lot = session.get(ProductoLote, link.id_lote)
            if lot:
                lot.cantidad_actual = int(lot.cantidad_actual or 0) + restore
                session.add(lot)
        link.cantidad_restaurada = int(link.cantidad_restaurada or 0) + restore
        session.add(link)
        remaining -= restore
    if remaining > 0:
        raise HTTPException(
            status_code=409,
            detail="No fue posible reconstruir todos los lotes de la venta.",
        )


def _update_movement_net(
    session: Session,
    Movimiento: Any,
    detail: VentaPOSDetalle,
    returned_inventory: int,
) -> None:
    link = session.exec(
        select(POSVentaMovimiento).where(
            POSVentaMovimiento.id_detalle == detail.id_detalle
        )
    ).first()
    if not link:
        return
    movement = session.get(Movimiento, link.id_movimiento)
    if not movement:
        return
    extra = _detail_extra(session, int(detail.id_detalle or 0))
    sold_inventory = int(extra.cantidad_inventario if extra else detail.cantidad or 0)
    net_inventory = max(sold_inventory - returned_inventory, 0)
    ratio = net_inventory / sold_inventory if sold_inventory else 0
    movement.cantidad = net_inventory
    movement.ingreso_total = _money(detail.subtotal * ratio)
    movement.costo_total = _money(detail.costo_total * ratio)
    movement.ganancia = _money(detail.ganancia * ratio)
    session.add(movement)


def _cash_summary(session: Session, cash_session: POSSesionCaja) -> Dict[str, Any]:
    controls = session.exec(
        select(POSVentaControl).where(POSVentaControl.id_sesion == cash_session.id_sesion)
    ).all()
    sale_ids = [row.id_venta for row in controls]
    sales = [session.get(VentaPOS, sale_id) for sale_id in sale_ids]
    sales = [row for row in sales if row is not None]
    active_sale_ids = [row.id_venta for row in sales if row.estado != "CANCELADA"]

    payments: List[VentaPOSPago] = []
    for sale_id in active_sale_ids:
        payments.extend(
            session.exec(
                select(VentaPOSPago).where(VentaPOSPago.id_venta == sale_id)
            ).all()
        )

    manual = session.exec(
        select(POSMovimientoEfectivo).where(
            POSMovimientoEfectivo.id_sesion == cash_session.id_sesion
        )
    ).all()
    returns = session.exec(
        select(POSDevolucion).where(POSDevolucion.id_sesion == cash_session.id_sesion)
    ).all()
    installments = session.exec(
        select(POSAbono).where(POSAbono.id_sesion == cash_session.id_sesion)
    ).all()

    payment_totals = defaultdict(float)
    for row in payments:
        payment_totals[row.metodo.lower()] += float(row.monto or 0)
    installment_totals = defaultdict(float)
    for row in installments:
        installment_totals[row.metodo.lower()] += float(row.monto or 0)

    entries = sum(
        row.monto
        for row in manual
        if row.tipo.upper() in {"ENTRADA", "AJUSTE_ENTRADA"}
    )
    exits = sum(
        row.monto
        for row in manual
        if row.tipo.upper() in {"RETIRO", "GASTO", "DEPOSITO", "AJUSTE_SALIDA"}
    )
    cash_refunds = 0.0
    for row in returns:
        if row.metodo_reembolso.lower() != "efectivo":
            continue
        extra = session.exec(
            select(POSDevolucionExtra).where(
                POSDevolucionExtra.id_devolucion == row.id_devolucion
            )
        ).first()
        cash_refunds += float(extra.reembolso_real if extra else row.monto or 0)

    expected = _money(
        cash_session.fondo_inicial
        + payment_totals["efectivo"]
        + installment_totals["efectivo"]
        + entries
        - exits
        - cash_refunds
    )
    return {
        "id_sesion": cash_session.id_sesion,
        "id_caja": cash_session.id_caja,
        "caja_nombre": cash_session.caja_nombre,
        "usuario": cash_session.usuario,
        "estado": cash_session.estado,
        "fondo_inicial": _money(cash_session.fondo_inicial),
        "fecha_apertura": cash_session.fecha_apertura,
        "fecha_cierre": cash_session.fecha_cierre,
        "ventas_completadas": len([s for s in sales if s.estado != "CANCELADA"]),
        "ventas_canceladas": len([s for s in sales if s.estado == "CANCELADA"]),
        "total_ventas": _money(sum(s.total for s in sales if s.estado != "CANCELADA")),
        "efectivo_ventas": _money(payment_totals["efectivo"]),
        "tarjeta": _money(payment_totals["tarjeta"]),
        "transferencia": _money(payment_totals["transferencia"]),
        "abonos_efectivo": _money(installment_totals["efectivo"]),
        "abonos_tarjeta": _money(installment_totals["tarjeta"]),
        "abonos_transferencia": _money(installment_totals["transferencia"]),
        "entradas_efectivo": _money(entries),
        "salidas_efectivo": _money(exits),
        "reembolsos_efectivo": _money(cash_refunds),
        "efectivo_esperado": expected,
        "efectivo_contado": (
            None if cash_session.efectivo_contado is None else _money(cash_session.efectivo_contado)
        ),
        "diferencia": None if cash_session.diferencia is None else _money(cash_session.diferencia),
        "observaciones": cash_session.observaciones,
        "movimientos_efectivo": [
            {
                "id_movimiento": row.id_movimiento,
                "tipo": row.tipo,
                "monto": _money(row.monto),
                "motivo": row.motivo,
                "usuario": row.usuario,
                "fecha": row.fecha,
            }
            for row in manual
        ],
    }



# RACKNOVA_RESUMEN_CAJA_EQUIPO
def _session_report(
    session: Session,
    cash_session: POSSesionCaja,
    Movimiento: Any,
) -> Dict[str, Any]:
    summary = _cash_summary(session, cash_session)

    controls = session.exec(
        select(POSVentaControl).where(
            POSVentaControl.id_sesion == cash_session.id_sesion
        )
    ).all()

    sale_ids = [
        int(row.id_venta)
        for row in controls
        if row.id_venta is not None
    ]

    sales: List[VentaPOS] = []

    if sale_ids:
        sales = session.exec(
            select(VentaPOS)
            .where(VentaPOS.id_venta.in_(sale_ids))
            .order_by(VentaPOS.fecha.asc())
        ).all()

    returns = session.exec(
        select(POSDevolucion)
        .where(POSDevolucion.id_sesion == cash_session.id_sesion)
        .order_by(POSDevolucion.fecha.asc())
    ).all()

    return_rows: List[Dict[str, Any]] = []

    for return_row in returns:
        details = session.exec(
            select(POSDevolucionDetalle).where(
                POSDevolucionDetalle.id_devolucion
                == return_row.id_devolucion
            )
        ).all()

        extra = session.exec(
            select(POSDevolucionExtra).where(
                POSDevolucionExtra.id_devolucion
                == return_row.id_devolucion
            )
        ).first()

        sale = session.get(VentaPOS, return_row.id_venta)

        return_rows.append(
            {
                "id_devolucion": return_row.id_devolucion,
                "folio": return_row.folio,
                "id_venta": return_row.id_venta,
                "folio_venta": sale.folio if sale else None,
                "usuario": return_row.usuario,
                "motivo": return_row.motivo,
                "metodo_reembolso": return_row.metodo_reembolso,
                "monto": _money(return_row.monto),
                "ajuste_credito": _money(
                    extra.ajuste_credito if extra else 0
                ),
                "reembolso_real": _money(
                    extra.reembolso_real
                    if extra
                    else return_row.monto
                ),
                "estado": return_row.estado,
                "fecha": return_row.fecha,
                "items": [
                    {
                        "id_detalle": item.id_detalle,
                        "sku": item.sku,
                        "nombre": item.nombre,
                        "cantidad": item.cantidad,
                        "precio_unitario": _money(
                            item.precio_unitario
                        ),
                        "subtotal": _money(item.subtotal),
                    }
                    for item in details
                ],
            }
        )

    product_movements: List[Dict[str, Any]] = []

    for sale in sales:
        details = session.exec(
            select(VentaPOSDetalle).where(
                VentaPOSDetalle.id_venta == sale.id_venta
            )
        ).all()

        for detail in details:
            detail_id = int(detail.id_detalle or 0)
            extra = _detail_extra(session, detail_id)
            factor = float(
                extra.factor_inventario if extra else 1
            ) or 1
            sold_quantity = float(
                extra.cantidad_venta
                if extra
                else detail.cantidad or 0
            )
            returned_inventory = _returned_inventory(
                session,
                detail_id,
            )
            returned_quantity = returned_inventory / factor
            net_quantity = max(
                sold_quantity - returned_quantity,
                0,
            )

            location = "Sin ubicación registrada"
            movement_link = session.exec(
                select(POSVentaMovimiento).where(
                    POSVentaMovimiento.id_detalle
                    == detail.id_detalle
                )
            ).first()

            if movement_link:
                movement = session.get(
                    Movimiento,
                    movement_link.id_movimiento,
                )

                if movement and getattr(
                    movement,
                    "ubicacion",
                    None,
                ):
                    location = str(movement.ubicacion)

            product_movements.append(
                {
                    "id_venta": sale.id_venta,
                    "id_detalle": detail.id_detalle,
                    "folio": sale.folio,
                    "fecha": sale.fecha,
                    "usuario": sale.usuario,
                    "estado_venta": sale.estado,
                    "sku": detail.sku,
                    "nombre": detail.nombre,
                    "ubicacion": location,
                    "unidad_venta": (
                        extra.unidad_venta
                        if extra
                        else "pieza"
                    ),
                    "cantidad_vendida": _qty(sold_quantity),
                    "cantidad_devuelta": _qty(
                        returned_quantity
                    ),
                    "cantidad_neta": _qty(net_quantity),
                    "precio_unitario": _money(
                        detail.precio_unitario_final
                    ),
                    "ingreso_neto": _money(
                        detail.precio_unitario_final
                        * net_quantity
                    ),
                }
            )

    valid_sales = [
        row
        for row in sales
        if row.estado != "CANCELADA"
    ]

    total_returns = _money(
        sum(row["monto"] for row in return_rows)
    )
    sales_total = _money(
        sum(row.total for row in valid_sales)
    )

    return {
        "sesion": summary,
        "periodo": {
            "inicio": cash_session.fecha_apertura,
            "fin": (
                cash_session.fecha_cierre
                or datetime.now()
            ),
        },
        "totales": {
            "ventas": sales_total,
            "devoluciones": total_returns,
            "ventas_netas": _money(
                sales_total - total_returns
            ),
            "numero_ventas": len(valid_sales),
            "ventas_canceladas": len(
                [
                    row
                    for row in sales
                    if row.estado == "CANCELADA"
                ]
            ),
            "numero_devoluciones": len(return_rows),
            "descuentos": _money(
                sum(row.descuento_total for row in valid_sales)
            ),
            "costo": _money(
                sum(row.costo_total for row in valid_sales)
            ),
            "ganancia_antes_devoluciones": _money(
                sum(row.ganancia for row in valid_sales)
            ),
        },
        "ventas": [
            _serialize_sale(session, row, True)
            for row in sales
        ],
        "devoluciones": return_rows,
        "movimientos_productos": product_movements,
        "movimientos_efectivo": summary.get(
            "movimientos_efectivo",
            [],
        ),
    }



# RACKNOVA_POS_V4_CAJAS_FIJAS
POS_V4_FIXED_BOXES = ((1, "Caja 1"), (2, "Caja 2"))


def _v4_model_fields(model: Any) -> Dict[str, Any]:
    return dict(
        getattr(model, "model_fields", None)
        or getattr(model, "__fields__", None)
        or {}
    )


def _v4_model_kwargs(model: Any, values: Dict[str, Any]) -> Dict[str, Any]:
    fields = _v4_model_fields(model)
    return {key: value for key, value in values.items() if key in fields}


def _v4_pk_name(model: Any, candidates: List[str]) -> Optional[str]:
    fields = _v4_model_fields(model)
    for name in candidates:
        if name in fields:
            return name
    return None


def _v4_user_keys(current_user: Any) -> set[str]:
    return {
        str(_key(current_user) or "").strip(),
        str(_name(current_user) or "").strip(),
    } - {""}


def _v4_session_belongs_to_user(
    cash_session: POSSesionCaja,
    current_user: Any,
) -> bool:
    return str(cash_session.usuario or "").strip() in _v4_user_keys(current_user)


def _v4_ensure_schema(session: Session) -> None:
    session.connection().execute(
        sa_text(
            """
            CREATE TABLE IF NOT EXISTS pos_mayoreo_menudeo (
                id_regla SERIAL PRIMARY KEY,
                sku VARCHAR(120) NOT NULL UNIQUE,
                nombre VARCHAR(255) NOT NULL,
                unidad VARCHAR(20) NOT NULL,
                precio_menudeo NUMERIC(14, 4) NOT NULL,
                cantidad_mayoreo NUMERIC(14, 4) NOT NULL,
                precio_mayoreo NUMERIC(14, 4) NOT NULL,
                cantidad_mayoreo_especial NUMERIC(14, 4),
                precio_mayoreo_especial NUMERIC(14, 4),
                fecha_inicio TIMESTAMP,
                fecha_fin TIMESTAMP,
                activo BOOLEAN NOT NULL DEFAULT TRUE,
                creado_en TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                actualizado_en TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    )
    session.commit()


def _v4_serialize_wholesale(row: Any) -> Dict[str, Any]:
    data = dict(row)
    for key in (
        "precio_menudeo",
        "cantidad_mayoreo",
        "precio_mayoreo",
        "cantidad_mayoreo_especial",
        "precio_mayoreo_especial",
    ):
        if data.get(key) is not None:
            data[key] = float(data[key])
    return data


def _v4_wholesale_price(
    session: Session,
    sku: str,
    unit: str,
    quantity: float,
    fallback_price: float,
) -> Dict[str, Any]:
    _v4_ensure_schema(session)
    row = session.connection().execute(
        sa_text(
            """
            SELECT *
            FROM pos_mayoreo_menudeo
            WHERE empresa_id = CAST(:empresa AS UUID)
              AND sku = :sku
              AND activo = TRUE
              AND unidad IN ('kg', 'litro')
              AND (fecha_inicio IS NULL OR fecha_inicio <= CURRENT_TIMESTAMP)
              AND (fecha_fin IS NULL OR fecha_fin >= CURRENT_TIMESTAMP)
            LIMIT 1
            """
        ),
        {"sku": sku, "empresa": str(session.info.get("racknova_empresa_id") or "11111111-1111-4111-8111-111111111111")},
    ).mappings().first()

    normalized_unit = str(unit or "").strip().lower()
    if not row or normalized_unit not in {"kg", "litro", "l"}:
        return {
            "precio": _money(fallback_price),
            "nivel": "normal",
            "regla": None,
        }

    quantity = max(float(quantity or 0), 0)
    price = float(row["precio_menudeo"])
    level = "menudeo"

    special_qty = row.get("cantidad_mayoreo_especial")
    special_price = row.get("precio_mayoreo_especial")

    if (
        special_qty is not None
        and special_price is not None
        and quantity >= float(special_qty)
    ):
        price = float(special_price)
        level = "mayoreo_especial"
    elif quantity >= float(row["cantidad_mayoreo"]):
        price = float(row["precio_mayoreo"])
        level = "mayoreo"

    return {
        "precio": _money(price),
        "nivel": level,
        "regla": _v4_serialize_wholesale(row),
    }


def _v4_ensure_fixed_box(
    session: Session,
    box_number: int,
) -> Dict[str, Any]:
    if int(box_number or 0) <= 0:
        raise HTTPException(status_code=400, detail="Identificador de caja inválido.")

    box_model = globals().get("POSCaja")
    if box_model is None:
        return {"id_caja": int(box_number), "nombre": f"Caja {int(box_number)}"}

    pk_field = _v4_pk_name(box_model, ["id_caja", "id"])
    name_field = _v4_pk_name(box_model, ["nombre", "caja_nombre", "name"])

    box = None
    if pk_field:
        try:
            box = session.get(box_model, int(box_number))
        except Exception:
            box = None

    name = f"Caja {int(box_number)}"
    if box is None and name_field:
        box = session.exec(
            select(box_model).where(getattr(box_model, name_field) == name)
        ).first()

    if box is None:
        values = {
            "nombre": name,
            "caja_nombre": name,
            "name": name,
            "activa": True,
            "activo": True,
            "estado": "DISPONIBLE",
            "numero": int(box_number),
        }
        box = box_model(**_v4_model_kwargs(box_model, values))
        session.add(box)
        session.commit()
        session.refresh(box)

    resolved_name = str(
        (getattr(box, name_field, None) if name_field else None)
        or name
    )
    return {
        "id_caja": getattr(box, pk_field, int(box_number)) if pk_field else int(box_number),
        "nombre": resolved_name,
    }


def _v4_open_sessions(session: Session) -> List[POSSesionCaja]:
    return list(
        session.exec(
            select(POSSesionCaja).where(POSSesionCaja.estado == "ABIERTA")
        ).all()
    )


def _v4_repair_session_links(
    session: Session,
    cash_session: POSSesionCaja,
) -> None:
    """Repara ventas/devoluciones antiguas sin id_sesion.

    Solo enlaza operaciones del mismo usuario y dentro del periodo de la caja.
    Esto corrige reportes vacíos creados antes de la actualización.
    """
    start = cash_session.fecha_apertura
    end = cash_session.fecha_cierre or datetime.now()
    user = str(cash_session.usuario or "").strip()

    linked_ids = {
        int(row.id_venta)
        for row in session.exec(select(POSVentaControl)).all()
        if getattr(row, "id_venta", None) is not None
    }

    sales_query = select(VentaPOS).where(
        VentaPOS.fecha >= start,
        VentaPOS.fecha <= end,
    )
    sales = session.exec(sales_query).all()

    for sale in sales:
        if str(getattr(sale, "usuario", "") or "").strip() != user:
            continue
        if int(sale.id_venta) in linked_ids:
            continue
        values = {
            "id_venta": sale.id_venta,
            "id_sesion": cash_session.id_sesion,
            "fecha": getattr(sale, "fecha", datetime.now()),
        }
        session.add(POSVentaControl(**_v4_model_kwargs(POSVentaControl, values)))
        linked_ids.add(int(sale.id_venta))

    returns = session.exec(
        select(POSDevolucion).where(
            POSDevolucion.fecha >= start,
            POSDevolucion.fecha <= end,
        )
    ).all()

    for row in returns:
        if str(getattr(row, "usuario", "") or "").strip() != user:
            continue
        if getattr(row, "id_sesion", None) in (None, 0):
            row.id_sesion = cash_session.id_sesion
            session.add(row)

    session.commit()


def _v4_safe_session_report(
    session: Session,
    cash_session: POSSesionCaja,
    Movimiento: Any,
) -> Dict[str, Any]:
    _v4_repair_session_links(session, cash_session)
    try:
        report = _session_report(session, cash_session, Movimiento)
        report["reporte_incompleto"] = False
        return report
    except Exception as exc:
        summary = _cash_summary(session, cash_session)
        return {
            "sesion": summary,
            "periodo": {
                "inicio": cash_session.fecha_apertura,
                "fin": cash_session.fecha_cierre or datetime.now(),
            },
            "totales": {
                "ventas": _money(getattr(cash_session, "total_ventas", 0)),
                "devoluciones": 0,
                "ventas_netas": _money(getattr(cash_session, "total_ventas", 0)),
                "numero_ventas": int(getattr(cash_session, "ventas_completadas", 0) or 0),
                "ventas_canceladas": 0,
                "numero_devoluciones": 0,
                "descuentos": 0,
                "costo": 0,
                "ganancia_antes_devoluciones": 0,
            },
            "ventas": [],
            "devoluciones": [],
            "movimientos_productos": [],
            "movimientos_efectivo": summary.get("movimientos_efectivo", []),
            "reporte_incompleto": True,
            "detalle_error": str(exc),
        }



# RACKNOVA_POS_V5_REPORTE_PROFESIONAL
def _v5_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _v5_text(value: Any, fallback: str = "") -> str:
    text_value = str(value or "").strip()
    return text_value or fallback


def _v5_sale_method(sale: Dict[str, Any]) -> str:
    raw = _v5_text(
        sale.get("metodo_pago")
        or sale.get("forma_pago")
        or sale.get("metodo")
        or sale.get("tipo_pago")
        or sale.get("payment_method"),
        "Otro",
    ).lower()

    if "efect" in raw:
        return "Efectivo"
    if "tarjet" in raw or "card" in raw:
        return "Tarjeta"
    if "transfer" in raw:
        return "Transferencia"
    if "crédit" in raw or "credit" in raw:
        return "Crédito"
    return "Otro"


def _v5_daily_report(
    session: Session,
    report_date: date,
    Movimiento: Any,
    cash_box: Optional[str] = None,
    operator: Optional[str] = None,
) -> Dict[str, Any]:
    start, end = _day_bounds(report_date)

    all_rows = list(
        session.exec(
            select(POSSesionCaja).where(
                POSSesionCaja.fecha_apertura < end,
                (POSSesionCaja.fecha_cierre == None)
                | (POSSesionCaja.fecha_cierre >= start),
            ).order_by(POSSesionCaja.fecha_apertura.asc())
        ).all()
    )

    available_boxes = sorted(
        {
            _v5_text(row.caja_nombre)
            for row in all_rows
            if _v5_text(row.caja_nombre)
        }
    )
    available_operators = sorted(
        {
            _v5_text(row.usuario)
            for row in all_rows
            if _v5_text(row.usuario)
        }
    )

    normalized_box = _v5_text(cash_box)
    normalized_operator = _v5_text(operator)

    rows = [
        row
        for row in all_rows
        if (
            not normalized_box
            or _v5_text(row.caja_nombre) == normalized_box
        )
        and (
            not normalized_operator
            or _v5_text(row.usuario) == normalized_operator
        )
    ]

    reports = [
        _v4_safe_session_report(session, row, Movimiento)
        for row in rows
    ]

    totals = {
        "ventas": 0.0,
        "devoluciones": 0.0,
        "ventas_netas": 0.0,
        "numero_ventas": 0,
        "ventas_canceladas": 0,
        "numero_devoluciones": 0,
        "descuentos": 0.0,
        "costo": 0.0,
        "ganancia": 0.0,
        "fondo_inicial": 0.0,
        "efectivo_esperado": 0.0,
        "efectivo_contado": 0.0,
        "diferencias": 0.0,
    }
    payment_methods: Dict[str, float] = {
        "Efectivo": 0.0,
        "Tarjeta": 0.0,
        "Transferencia": 0.0,
        "Crédito": 0.0,
        "Otro": 0.0,
    }
    by_box: Dict[str, Dict[str, Any]] = {}
    by_operator: Dict[str, Dict[str, Any]] = {}
    products: Dict[tuple[str, str], Dict[str, Any]] = {}
    returns: List[Dict[str, Any]] = []
    cash_movements: List[Dict[str, Any]] = []

    def grouping_row(container: Dict[str, Dict[str, Any]], key: str) -> Dict[str, Any]:
        return container.setdefault(
            key,
            {
                "nombre": key,
                "sesiones": 0,
                "ventas": 0.0,
                "devoluciones": 0.0,
                "venta_neta": 0.0,
                "operaciones": 0,
                "diferencia": 0.0,
            },
        )

    for report in reports:
        report_totals = report.get("totales") or {}
        cash_session = report.get("sesion") or {}
        box_name = _v5_text(cash_session.get("caja_nombre"), "Sin caja")
        operator_name = _v5_text(cash_session.get("usuario"), "Sin operador")

        totals["ventas"] += _v5_number(report_totals.get("ventas"))
        totals["devoluciones"] += _v5_number(report_totals.get("devoluciones"))
        totals["ventas_netas"] += _v5_number(report_totals.get("ventas_netas"))
        totals["numero_ventas"] += int(report_totals.get("numero_ventas") or 0)
        totals["ventas_canceladas"] += int(report_totals.get("ventas_canceladas") or 0)
        totals["numero_devoluciones"] += int(report_totals.get("numero_devoluciones") or 0)
        totals["descuentos"] += _v5_number(report_totals.get("descuentos"))
        totals["costo"] += _v5_number(report_totals.get("costo"))
        totals["ganancia"] += _v5_number(
            report_totals.get("ganancia_antes_devoluciones")
        )
        totals["fondo_inicial"] += _v5_number(cash_session.get("fondo_inicial"))
        totals["efectivo_esperado"] += _v5_number(cash_session.get("efectivo_esperado"))
        totals["efectivo_contado"] += _v5_number(cash_session.get("efectivo_contado"))
        totals["diferencias"] += _v5_number(cash_session.get("diferencia"))

        box_row = grouping_row(by_box, box_name)
        operator_row = grouping_row(by_operator, operator_name)
        for target in (box_row, operator_row):
            target["sesiones"] += 1
            target["ventas"] += _v5_number(report_totals.get("ventas"))
            target["devoluciones"] += _v5_number(report_totals.get("devoluciones"))
            target["venta_neta"] += _v5_number(report_totals.get("ventas_netas"))
            target["operaciones"] += int(report_totals.get("numero_ventas") or 0)
            target["diferencia"] += _v5_number(cash_session.get("diferencia"))

        session_method_values = {
            "Efectivo": _v5_number(cash_session.get("efectivo_ventas")),
            "Tarjeta": _v5_number(cash_session.get("tarjeta")),
            "Transferencia": _v5_number(cash_session.get("transferencia")),
            "Crédito": _v5_number(
                cash_session.get("credito")
                or cash_session.get("ventas_credito")
            ),
        }

        identified = sum(session_method_values.values())
        if identified > 0:
            for method_name, amount in session_method_values.items():
                payment_methods[method_name] += amount
            payment_methods["Otro"] += max(
                _v5_number(report_totals.get("ventas")) - identified,
                0,
            )
        else:
            for sale in report.get("ventas") or []:
                if _v5_text(sale.get("estado")).upper() == "CANCELADA":
                    continue
                method_name = _v5_sale_method(sale)
                payment_methods[method_name] += _v5_number(
                    sale.get("total")
                    or sale.get("total_venta")
                    or sale.get("monto")
                )

        for movement in report.get("movimientos_productos") or []:
            sku = _v5_text(movement.get("sku"), "SIN-SKU")
            unit = _v5_text(movement.get("unidad_venta"), "pieza")
            key = (sku, unit)
            item = products.setdefault(
                key,
                {
                    "sku": sku,
                    "nombre": _v5_text(movement.get("nombre"), "Producto"),
                    "unidad": unit,
                    "cantidad_vendida": 0.0,
                    "cantidad_devuelta": 0.0,
                    "cantidad_neta": 0.0,
                    "ingreso_neto": 0.0,
                },
            )
            item["cantidad_vendida"] += _v5_number(movement.get("cantidad_vendida"))
            item["cantidad_devuelta"] += _v5_number(movement.get("cantidad_devuelta"))
            item["cantidad_neta"] += _v5_number(movement.get("cantidad_neta"))
            item["ingreso_neto"] += _v5_number(movement.get("ingreso_neto"))

        for return_item in report.get("devoluciones") or []:
            returns.append(
                {
                    **return_item,
                    "caja": box_name,
                    "operador_caja": operator_name,
                    "id_sesion": cash_session.get("id_sesion"),
                }
            )

        for cash_movement in report.get("movimientos_efectivo") or []:
            cash_movements.append(
                {
                    **cash_movement,
                    "caja": box_name,
                    "operador_caja": operator_name,
                    "id_sesion": cash_session.get("id_sesion"),
                }
            )

    for value in totals:
        if value not in {
            "numero_ventas",
            "ventas_canceladas",
            "numero_devoluciones",
        }:
            totals[value] = _money(totals[value])

    for collection in (by_box, by_operator):
        for item in collection.values():
            for key in ("ventas", "devoluciones", "venta_neta", "diferencia"):
                item[key] = _money(item[key])

    product_rows = []
    for item in products.values():
        for key in (
            "cantidad_vendida",
            "cantidad_devuelta",
            "cantidad_neta",
        ):
            item[key] = _qty(item[key])
        item["ingreso_neto"] = _money(item["ingreso_neto"])
        product_rows.append(item)

    product_rows.sort(
        key=lambda item: (
            -_v5_number(item.get("ingreso_neto")),
            _v5_text(item.get("nombre")),
        )
    )

    payment_rows = [
        {"metodo": name, "monto": _money(amount)}
        for name, amount in payment_methods.items()
        if abs(amount) > 0.0001
    ]

    margin = (
        (totals["ganancia"] / totals["ventas_netas"] * 100)
        if totals["ventas_netas"]
        else 0
    )
    totals["margen_porcentaje"] = round(margin, 2)

    return {
        "fecha": report_date.isoformat(),
        "generado_en": datetime.now(),
        "filtros": {
            "caja": normalized_box or None,
            "operador": normalized_operator or None,
        },
        "catalogos": {
            "cajas": available_boxes,
            "operadores": available_operators,
        },
        "totales": totals,
        "metodos_pago": payment_rows,
        "cajas": sorted(by_box.values(), key=lambda item: item["nombre"]),
        "operadores": sorted(by_operator.values(), key=lambda item: item["nombre"]),
        "productos": product_rows,
        "devoluciones": sorted(
            returns,
            key=lambda item: _v5_text(item.get("fecha")),
        ),
        "movimientos_efectivo": sorted(
            cash_movements,
            key=lambda item: _v5_text(item.get("fecha")),
        ),
        "sesiones": reports,
    }



# ==========================================================
# RACKNOVA_MULTIEMPRESA_FASE1
# Base multiempresa: empresas, membresías y contexto seguro.
# ==========================================================
MULTIEMPRESA_DEFAULT_ID = "11111111-1111-4111-8111-111111111111"


def _multi_user_keys(current_user: Any) -> List[str]:
    values: List[str] = []
    for value in (
        _key(current_user),
        _name(current_user),
        getattr(current_user, "email", None),
        getattr(current_user, "username", None),
        getattr(current_user, "id", None),
        getattr(current_user, "id_usuario", None),
    ):
        normalized = str(value or "").strip()
        if normalized and normalized not in values:
            values.append(normalized)
    return values


def _multi_primary_user_key(current_user: Any) -> str:
    keys = _multi_user_keys(current_user)
    if not keys:
        raise HTTPException(
            status_code=401,
            detail="No se pudo identificar al usuario autenticado.",
        )
    return keys[0]


def _multi_normalize_role(current_user: Any, owner_for_admin: bool = False) -> str:
    role = str(_role(current_user) or "viewer").strip().lower()
    if role == "admin" and owner_for_admin:
        return "owner"
    if role in {"owner", "admin", "operator", "viewer"}:
        return role
    return "viewer"


def _multi_slug(value: Any) -> str:
    raw = str(value or "").strip().lower()
    result: List[str] = []
    dash = False
    for char in raw:
        if char.isalnum():
            result.append(char)
            dash = False
        elif result and not dash:
            result.append("-")
            dash = True
    return "".join(result).strip("-")[:120]


def _multi_ensure_schema(session: Session) -> None:
    # Esquema mínimo de respaldo. La migración SQL de Supabase sigue siendo
    # la vía recomendada porque además tenantiza las tablas existentes.
    session.connection().execute(
        sa_text(
            """
            CREATE TABLE IF NOT EXISTS empresas (
                id_empresa UUID PRIMARY KEY,
                nombre VARCHAR(150) NOT NULL,
                slug VARCHAR(120) NOT NULL UNIQUE,
                activo BOOLEAN NOT NULL DEFAULT TRUE,
                plan VARCHAR(30) NOT NULL DEFAULT 'basic',
                moneda VARCHAR(10) NOT NULL DEFAULT 'MXN',
                zona_horaria VARCHAR(80) NOT NULL DEFAULT 'America/Mexico_City',
                rfc VARCHAR(20),
                razon_social VARCHAR(255),
                logo_url TEXT,
                proveedor_facturacion VARCHAR(50),
                proveedor_organizacion_id VARCHAR(255),
                creado_en TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                actualizado_en TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    session.connection().execute(
        sa_text(
            """
            CREATE TABLE IF NOT EXISTS empresa_usuarios (
                id_membresia UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                id_empresa UUID NOT NULL REFERENCES empresas(id_empresa) ON DELETE CASCADE,
                usuario_key VARCHAR(255) NOT NULL,
                nombre_usuario VARCHAR(255),
                rol VARCHAR(30) NOT NULL DEFAULT 'viewer',
                activo BOOLEAN NOT NULL DEFAULT TRUE,
                creado_en TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                actualizado_en TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                UNIQUE (id_empresa, usuario_key)
            )
            """
        )
    )
    session.connection().execute(
        sa_text(
            """
            INSERT INTO empresas (
                id_empresa, nombre, slug, activo, plan, moneda, zona_horaria
            ) VALUES (
                CAST(:id AS UUID),
                'RackNova Principal',
                'racknova-principal',
                TRUE,
                'legacy',
                'MXN',
                'America/Mexico_City'
            )
            ON CONFLICT (id_empresa) DO NOTHING
            """
        ),
        {"id": MULTIEMPRESA_DEFAULT_ID},
    )
    session.commit()


def _multi_ensure_legacy_membership(
    session: Session,
    current_user: Any,
) -> None:
    _multi_ensure_schema(session)
    keys = _multi_user_keys(current_user)
    if not keys:
        return

    existing = session.connection().execute(
        sa_text(
            """
            SELECT 1
            FROM empresa_usuarios
            WHERE usuario_key = ANY(:keys)
              AND activo = TRUE
            LIMIT 1
            """
        ),
        {"keys": keys},
    ).first()
    if existing:
        return

    primary = keys[0]
    session.connection().execute(
        sa_text(
            """
            INSERT INTO empresa_usuarios (
                id_empresa,
                usuario_key,
                nombre_usuario,
                rol,
                activo
            ) VALUES (
                CAST(:empresa AS UUID),
                :usuario,
                :nombre,
                :rol,
                TRUE
            )
            ON CONFLICT (id_empresa, usuario_key) DO NOTHING
            """
        ),
        {
            "empresa": MULTIEMPRESA_DEFAULT_ID,
            "usuario": primary,
            "nombre": str(_name(current_user) or primary),
            "rol": _multi_normalize_role(current_user, owner_for_admin=True),
        },
    )
    session.commit()


def _multi_memberships(
    session: Session,
    current_user: Any,
) -> List[Dict[str, Any]]:
    _multi_ensure_legacy_membership(session, current_user)
    keys = _multi_user_keys(current_user)
    rows = session.connection().execute(
        sa_text(
            """
            SELECT
                e.id_empresa,
                e.nombre,
                e.slug,
                e.activo,
                e.plan,
                e.moneda,
                e.zona_horaria,
                eu.rol,
                eu.usuario_key
            FROM empresa_usuarios eu
            JOIN empresas e ON e.id_empresa = eu.id_empresa
            WHERE eu.usuario_key = ANY(:keys)
              AND eu.activo = TRUE
              AND e.activo = TRUE
            ORDER BY
                CASE WHEN e.id_empresa = CAST(:principal AS UUID) THEN 0 ELSE 1 END,
                e.nombre
            """
        ),
        {"keys": keys, "principal": MULTIEMPRESA_DEFAULT_ID},
    ).mappings().all()
    return [dict(row) for row in rows]


def _multi_empresa_context(
    session: Session,
    current_user: Any,
    empresa_id: Optional[str] = None,
    allowed_roles: Optional[set[str]] = None,
) -> Dict[str, Any]:
    memberships = _multi_memberships(session, current_user)
    if not memberships:
        raise HTTPException(
            status_code=403,
            detail="El usuario no pertenece a ninguna empresa activa.",
        )

    requested = str(empresa_id or "").strip()
    selected: Optional[Dict[str, Any]] = None

    if requested:
        for row in memberships:
            if str(row.get("id_empresa")) == requested:
                selected = row
                break
        if selected is None:
            raise HTTPException(
                status_code=403,
                detail="No tienes acceso a la empresa solicitada.",
            )
    else:
        selected = memberships[0]

    role = str(selected.get("rol") or "viewer").lower()
    if allowed_roles and role not in allowed_roles:
        raise HTTPException(
            status_code=403,
            detail="Tu rol no permite realizar esta acción en esta empresa.",
        )

    return selected


def _multi_create_company(
    session: Session,
    current_user: Any,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    rn_tenant.require_platform_superadmin(session, current_user)
    _multi_ensure_legacy_membership(session, current_user)

    name = str(payload.get("nombre") or "").strip()
    if len(name) < 2:
        raise HTTPException(status_code=400, detail="Escribe el nombre de la empresa.")

    slug = _multi_slug(payload.get("slug") or name)
    if not slug:
        raise HTTPException(status_code=400, detail="No se pudo generar el identificador de empresa.")

    exists = session.connection().execute(
        sa_text("SELECT 1 FROM empresas WHERE slug = :slug LIMIT 1"),
        {"slug": slug},
    ).first()
    if exists:
        raise HTTPException(status_code=409, detail="Ya existe una empresa con ese identificador.")

    row = session.connection().execute(
        sa_text(
            """
            INSERT INTO empresas (
                id_empresa,
                nombre,
                slug,
                activo,
                plan,
                moneda,
                zona_horaria
            ) VALUES (
                gen_random_uuid(),
                :nombre,
                :slug,
                TRUE,
                :plan,
                :moneda,
                :zona_horaria
            )
            RETURNING id_empresa, nombre, slug, activo, plan, moneda, zona_horaria
            """
        ),
        {
            "nombre": name,
            "slug": slug,
            "plan": str(payload.get("plan") or "basic").strip() or "basic",
            "moneda": str(payload.get("moneda") or "MXN").strip() or "MXN",
            "zona_horaria": str(
                payload.get("zona_horaria") or "America/Mexico_City"
            ).strip()
            or "America/Mexico_City",
        },
    ).mappings().first()

    user_key = _multi_primary_user_key(current_user)
    session.connection().execute(
        sa_text(
            """
            INSERT INTO empresa_usuarios (
                id_empresa,
                usuario_key,
                nombre_usuario,
                rol,
                activo
            ) VALUES (
                CAST(:empresa AS UUID),
                :usuario,
                :nombre,
                'owner',
                TRUE
            )
            ON CONFLICT (id_empresa, usuario_key)
            DO UPDATE SET rol = 'owner', activo = TRUE, actualizado_en = NOW()
            """
        ),
        {
            "empresa": str(row["id_empresa"]),
            "usuario": user_key,
            "nombre": str(_name(current_user) or user_key),
        },
    )
    session.commit()
    return dict(row)


def _multi_add_member(
    session: Session,
    current_user: Any,
    empresa_id: str,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    _multi_empresa_context(
        session,
        current_user,
        empresa_id,
        {"owner", "admin"},
    )

    usuario_key = str(
        payload.get("usuario_key")
        or payload.get("email")
        or payload.get("usuario")
        or ""
    ).strip()
    if not usuario_key:
        raise HTTPException(
            status_code=400,
            detail="Indica usuario_key o email del usuario.",
        )

    role = str(payload.get("rol") or "viewer").strip().lower()
    if role not in {"admin", "operator", "viewer"}:
        raise HTTPException(
            status_code=400,
            detail="Rol inválido. Usa admin, operator o viewer.",
        )

    session.connection().execute(
        sa_text(
            """
            INSERT INTO empresa_usuarios (
                id_empresa,
                usuario_key,
                nombre_usuario,
                rol,
                activo
            ) VALUES (
                CAST(:empresa AS UUID),
                :usuario,
                :nombre,
                :rol,
                TRUE
            )
            ON CONFLICT (id_empresa, usuario_key)
            DO UPDATE SET
                nombre_usuario = EXCLUDED.nombre_usuario,
                rol = EXCLUDED.rol,
                activo = TRUE,
                actualizado_en = NOW()
            """
        ),
        {
            "empresa": empresa_id,
            "usuario": usuario_key,
            "nombre": str(payload.get("nombre_usuario") or usuario_key),
            "rol": role,
        },
    )
    session.commit()
    rn_tenant.raise_global_role_if_needed(session, usuario_key, role)
    return {
        "id_empresa": empresa_id,
        "usuario_key": usuario_key,
        "rol": role,
        "activo": True,
    }


def _multi_diagnostic(session: Session) -> Dict[str, Any]:
    _multi_ensure_schema(session)
    company_count = session.connection().execute(
        sa_text("SELECT COUNT(*) FROM empresas")
    ).scalar_one()
    membership_count = session.connection().execute(
        sa_text("SELECT COUNT(*) FROM empresa_usuarios")
    ).scalar_one()

    tables = session.connection().execute(
        sa_text(
            """
            SELECT
                t.table_name,
                EXISTS (
                    SELECT 1
                    FROM information_schema.columns c
                    WHERE c.table_schema = t.table_schema
                      AND c.table_name = t.table_name
                      AND c.column_name = 'empresa_id'
                ) AS tiene_empresa_id
            FROM information_schema.tables t
            WHERE t.table_schema = 'public'
              AND t.table_type = 'BASE TABLE'
            ORDER BY t.table_name
            """
        )
    ).mappings().all()

    tenant_candidates: List[Dict[str, Any]] = []
    prefixes = (
        "pos_", "producto", "inventario", "movimiento", "venta",
        "cliente", "credito", "crédito", "abono", "cotizacion",
        "cotización", "catalogo", "catálogo", "ubicacion", "ubicación",
        "rack", "alerta",
    )
    for row in tables:
        name = str(row["table_name"])
        if name.startswith(prefixes):
            tenant_candidates.append(dict(row))

    missing = [
        row["table_name"]
        for row in tenant_candidates
        if not bool(row["tiene_empresa_id"])
    ]

    return {
        "fase": 1,
        "empresa_principal": MULTIEMPRESA_DEFAULT_ID,
        "empresas": int(company_count),
        "membresias": int(membership_count),
        "tablas_comerciales_detectadas": tenant_candidates,
        "tablas_pendientes_empresa_id": missing,
        "selector_dashboard_habilitado": False,
        "motivo_selector_deshabilitado": (
            "Se habilitará cuando los módulos operativos filtren por empresa."
        ),
    }


def _day_bounds(report_date: date) -> Tuple[datetime, datetime]:
    return datetime.combine(report_date, time.min), datetime.combine(report_date, time.max)


def _daily_report(
    session: Session,
    report_date: date,
    Movimiento: Any = None,
) -> Dict[str, Any]:
    start, end = _day_bounds(report_date)
    sales = session.exec(
        select(VentaPOS).where((VentaPOS.fecha >= start) & (VentaPOS.fecha <= end))
    ).all()
    returns = session.exec(
        select(POSDevolucion).where(
            (POSDevolucion.fecha >= start) & (POSDevolucion.fecha <= end)
        )
    ).all()
    installments = session.exec(
        select(POSAbono).where((POSAbono.fecha >= start) & (POSAbono.fecha <= end))
    ).all()
    cash_sessions = session.exec(
        select(POSSesionCaja).where(
            (POSSesionCaja.fecha_apertura >= start)
            & (POSSesionCaja.fecha_apertura <= end)
        )
    ).all()

    active_sales = [row for row in sales if row.estado != "CANCELADA"]
    cancelled = [row for row in sales if row.estado == "CANCELADA"]
    return_total = _money(sum(row.monto for row in returns))
    gross = _money(sum(row.subtotal for row in active_sales))
    discounts = _money(sum(row.descuento_total for row in active_sales))
    net_before_returns = _money(sum(row.total for row in active_sales))
    net = _money(net_before_returns - return_total)
    costs = _money(sum(row.costo_total for row in active_sales))
    # El costo retornado se aproxima desde detalle de devolución.
    returned_cost = 0.0
    product_map: Dict[str, Dict[str, Any]] = {}
    movement_details: List[Dict[str, Any]] = []
    cashier_map = defaultdict(lambda: {"ventas": 0, "total": 0.0})
    box_map = defaultdict(lambda: {"ventas": 0, "total": 0.0})
    hourly = defaultdict(float)
    payment_map = defaultdict(float)

    active_ids = {int(row.id_venta or 0) for row in active_sales}
    for sale in active_sales:
        cashier_map[sale.usuario]["ventas"] += 1
        cashier_map[sale.usuario]["total"] += float(sale.total or 0)
        hourly[f"{sale.fecha.hour:02d}:00"] += float(sale.total or 0)
        control = _sale_control(session, int(sale.id_venta or 0))
        box_name = "Sin caja"
        if control:
            cash_session = session.get(POSSesionCaja, control.id_sesion)
            if cash_session:
                box_name = cash_session.caja_nombre
        box_map[box_name]["ventas"] += 1
        box_map[box_name]["total"] += float(sale.total or 0)

        for payment in session.exec(
            select(VentaPOSPago).where(VentaPOSPago.id_venta == sale.id_venta)
        ).all():
            payment_map[payment.metodo.lower()] += float(payment.monto or 0)

        for detail in session.exec(
            select(VentaPOSDetalle).where(VentaPOSDetalle.id_venta == sale.id_venta)
        ).all():
            extra = _detail_extra(session, int(detail.id_detalle or 0))
            factor = float(extra.factor_inventario if extra else 1) or 1
            sold_qty = float(extra.cantidad_venta if extra else detail.cantidad or 0)
            returned_inventory = _returned_inventory(session, int(detail.id_detalle or 0))
            returned_qty = returned_inventory / factor
            net_qty = max(sold_qty - returned_qty, 0)
            ratio = returned_qty / sold_qty if sold_qty else 0

            movement_location = "Sin ubicación registrada"
            if Movimiento is not None and detail.id_detalle is not None:
                movement_link = session.exec(
                    select(POSVentaMovimiento).where(
                        POSVentaMovimiento.id_detalle == detail.id_detalle
                    )
                ).first()
                if movement_link:
                    movement_row = session.get(Movimiento, movement_link.id_movimiento)
                    if movement_row and getattr(movement_row, "ubicacion", None):
                        movement_location = str(movement_row.ubicacion)

            returned_cost += float(detail.costo_total or 0) * ratio
            row = product_map.setdefault(
                detail.sku,
                {
                    "sku": detail.sku,
                    "nombre": detail.nombre,
                    "unidad_venta": extra.unidad_venta if extra else "pieza",
                    "cantidad": 0.0,
                    "ingresos": 0.0,
                    "ganancia": 0.0,
                },
            )
            row["cantidad"] += net_qty
            row["ingresos"] += float(detail.subtotal or 0) * (1 - ratio)
            row["ganancia"] += float(detail.ganancia or 0) * (1 - ratio)

            movement_details.append(
                {
                    "id_venta": int(sale.id_venta or 0),
                    "id_detalle": int(detail.id_detalle or 0),
                    "folio": sale.folio,
                    "fecha": sale.fecha,
                    "usuario": sale.usuario,
                    "caja": box_name,
                    "sku": detail.sku,
                    "nombre": detail.nombre,
                    "ubicacion": movement_location,
                    "unidad_venta": extra.unidad_venta if extra else "pieza",
                    "cantidad_vendida": _qty(sold_qty),
                    "cantidad_devuelta": _qty(returned_qty),
                    "cantidad_neta": _qty(net_qty),
                    "ingresos": _money(float(detail.subtotal or 0) * (1 - ratio)),
                    "ganancia": _money(float(detail.ganancia or 0) * (1 - ratio)),
                }
            )

    for installment in installments:
        payment_map[f"abono_{installment.metodo.lower()}"] += float(installment.monto or 0)

    cost_net = _money(costs - returned_cost)
    profit = _money(net - cost_net)
    margin = round((profit / net) * 100, 2) if net > 0 else 0
    top_products = sorted(
        (
            {
                **row,
                "cantidad": _qty(row["cantidad"]),
                "ingresos": _money(row["ingresos"]),
                "ganancia": _money(row["ganancia"]),
            }
            for row in product_map.values()
        ),
        key=lambda item: item["ingresos"],
        reverse=True,
    )

    cash_summaries = [_cash_summary(session, row) for row in cash_sessions]
    return {
        "fecha": report_date,
        "generado_en": datetime.now(),
        "resumen": {
            "numero_ventas": len(active_sales),
            "ventas_canceladas": len(cancelled),
            "devoluciones": len(returns),
            "ventas_brutas": gross,
            "descuentos": discounts,
            "ventas_antes_devoluciones": net_before_returns,
            "monto_devoluciones": return_total,
            "ventas_netas": net,
            "costo_mercancia": cost_net,
            "ganancia": profit,
            "margen": margin,
            "abonos": _money(sum(row.monto for row in installments)),
        },
        "metodos_pago": {key: _money(value) for key, value in payment_map.items()},
        "productos": top_products,
        "movimientos_productos": sorted(
            movement_details,
            key=lambda item: (
                str(item["fecha"]),
                item["ubicacion"],
                item["sku"],
            ),
        ),
        "cajeros": [
            {"usuario": key, "ventas": value["ventas"], "total": _money(value["total"])}
            for key, value in sorted(cashier_map.items())
        ],
        "cajas": [
            {"caja": key, "ventas": value["ventas"], "total": _money(value["total"])}
            for key, value in sorted(box_map.items())
        ],
        "ventas_por_hora": [
            {"hora": key, "total": _money(value)}
            for key, value in sorted(hourly.items())
        ],
        "cortes": cash_summaries,
    }


def _xlsx_report(data: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Falta instalar openpyxl en requirements.txt.",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen diario"
    sheet["A1"] = "RackNova — Reporte diario"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet["A2"] = str(data["fecha"])
    row = 4
    for key, value in data["resumen"].items():
        sheet.cell(row=row, column=1, value=key.replace("_", " ").title())
        sheet.cell(row=row, column=2, value=value)
        row += 1

    products = workbook.create_sheet("Productos")
    headers = ["SKU", "Producto", "Unidad", "Cantidad", "Ingresos", "Ganancia"]
    products.append(headers)
    for cell in products[1]:
        cell.font = Font(bold=True)
    for item in data["productos"]:
        products.append(
            [
                item["sku"],
                item["nombre"],
                item["unidad_venta"],
                item["cantidad"],
                item["ingresos"],
                item["ganancia"],
            ]
        )

    movements = workbook.create_sheet("Salidas y ubicaciones")
    movement_headers = [
        "Fecha",
        "Folio",
        "Caja",
        "Cajero",
        "SKU",
        "Producto",
        "Ubicación",
        "Unidad",
        "Cantidad vendida",
        "Cantidad devuelta",
        "Cantidad neta",
        "Ingresos netos",
        "Ganancia",
    ]
    movements.append(movement_headers)
    for cell in movements[1]:
        cell.font = Font(bold=True)
    for item in data.get("movimientos_productos", []):
        movements.append(
            [
                str(item["fecha"]),
                item["folio"],
                item["caja"],
                item["usuario"],
                item["sku"],
                item["nombre"],
                item["ubicacion"],
                item["unidad_venta"],
                item["cantidad_vendida"],
                item["cantidad_devuelta"],
                item["cantidad_neta"],
                item["ingresos"],
                item["ganancia"],
            ]
        )
    movements.freeze_panes = "A2"
    movements.auto_filter.ref = movements.dimensions
    movement_widths = {
        "A": 20,
        "B": 18,
        "C": 18,
        "D": 22,
        "E": 16,
        "F": 32,
        "G": 20,
        "H": 14,
        "I": 18,
        "J": 18,
        "K": 16,
        "L": 16,
        "M": 16,
    }
    for column, width in movement_widths.items():
        movements.column_dimensions[column].width = width

    cuts = workbook.create_sheet("Cortes de caja")
    cuts.append(
        [
            "Caja",
            "Usuario",
            "Apertura",
            "Cierre",
            "Esperado",
            "Contado",
            "Diferencia",
        ]
    )
    for cell in cuts[1]:
        cell.font = Font(bold=True)
    for item in data["cortes"]:
        cuts.append(
            [
                item["caja_nombre"],
                item["usuario"],
                str(item["fecha_apertura"]),
                str(item["fecha_cierre"] or ""),
                item["efectivo_esperado"],
                item["efectivo_contado"],
                item["diferencia"],
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _pdf_report(data: Dict[str, Any]) -> bytes:
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError as exc:
        raise HTTPException(
            status_code=500,
            detail="Falta instalar reportlab en requirements.txt.",
        ) from exc

    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=letter)
    width, height = letter
    y = height - 50
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(45, y, "RackNova - Reporte diario de ventas")
    y -= 24
    pdf.setFont("Helvetica", 11)
    pdf.drawString(45, y, f"Fecha: {data['fecha']}")
    y -= 30
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(45, y, "Resumen")
    y -= 18
    pdf.setFont("Helvetica", 10)
    for key, value in data["resumen"].items():
        label = key.replace("_", " ").title()
        pdf.drawString(55, y, f"{label}: {value}")
        y -= 15
        if y < 80:
            pdf.showPage()
            y = height - 50
    y -= 10
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(45, y, "Productos principales")
    y -= 18
    pdf.setFont("Helvetica", 9)
    for item in data["productos"][:20]:
        text = (
            f"{item['sku']} - {item['nombre']}: {item['cantidad']} "
            f"{item['unidad_venta']} | ${item['ingresos']:.2f}"
        )
        pdf.drawString(55, y, text[:105])
        y -= 14
        if y < 70:
            pdf.showPage()
            y = height - 50
    if y < 110:
        pdf.showPage()
        y = height - 50

    y -= 12
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(45, y, "Conciliacion fisica de salidas")
    y -= 18
    pdf.setFont("Helvetica", 7.5)

    for item in data.get("movimientos_productos", []):
        fecha_value = str(item.get("fecha", ""))
        hora = fecha_value[11:16] if len(fecha_value) >= 16 else fecha_value
        text = (
            f"{hora} | {item['folio']} | {item['ubicacion']} | "
            f"{item['sku']} - {item['nombre']} | "
            f"Neto: {item['cantidad_neta']} {item['unidad_venta']}"
        )
        pdf.drawString(45, y, text[:135])
        y -= 12
        if y < 55:
            pdf.showPage()
            y = height - 45
            pdf.setFont("Helvetica", 7.5)

    pdf.save()
    return output.getvalue()


# ==========================================================
# REGISTRO DE ENDPOINTS
# ==========================================================


def registrar_modulo_pos_fase3(
    *,
    app: Any,
    get_session: Callable[..., Any],
    require_roles: Callable[..., Any],
    Producto: Any,
    ProductoLote: Any,
    Movimiento: Any,
    mexico_now: Callable[[], datetime],
    descontar_lotes_fefo: Callable[..., List[Dict[str, Any]]],
    obtener_caducidad_mas_proxima: Callable[..., Any],
) -> None:
    read_user = require_roles("admin", "operator", "viewer")
    operator_user = require_roles("admin", "operator")
    admin_user = require_roles("admin")

    @app.get("/pos/v3/estado")
    def estado_v3(current_user: Any = Depends(read_user)):
        return {
            "version": 3,
            "disponible": True,
            "modulos": [
                "reportes",
                "clientes",
                "credito",
                "abonos",
                "promociones",
                "precios_avanzados",
                "productos_fraccionados",
                "auditoria",
            ],
        }

    # -------------------- CLIENTES --------------------

    @app.get("/pos/v3/clientes")
    def list_clients(
        query: str = Query(default="", max_length=150),
        include_inactive: bool = Query(default=False),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        statement = select(POSCliente).order_by(POSCliente.nombre.asc())
        if not include_inactive:
            statement = statement.where(POSCliente.activo == True)  # noqa: E712
        text_value = query.strip()
        if text_value:
            pattern = f"%{text_value}%"
            statement = statement.where(
                or_(
                    POSCliente.nombre.ilike(pattern),
                    POSCliente.telefono.ilike(pattern),
                    POSCliente.email.ilike(pattern),
                    POSCliente.rfc.ilike(pattern),
                )
            )
        rows = session.exec(statement).all()
        return _serialize_clients_bulk(rows, session)

    @app.post("/pos/v3/clientes")
    def create_client(
        data: ClienteRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        now = mexico_now()
        row = POSCliente(
            **{**_dump(data), "nombre": data.nombre.strip(), "telefono": data.telefono.strip() if data.telefono else None, "email": data.email.strip().lower() if data.email else None, "rfc": data.rfc.strip().upper() if data.rfc else None},
            fecha_creacion=now,
            fecha_actualizacion=now,
        )
        session.add(row)
        session.flush()
        _audit(
            session,
            action="CREAR_CLIENTE",
            entity="cliente",
            entity_id=row.id_cliente,
            details={"nombre": row.nombre},
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)
        return _serialize_client(row, session)

    @app.put("/pos/v3/clientes/{client_id}")
    def update_client(
        client_id: int,
        data: ClienteRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(POSCliente, client_id)
        if not row:
            raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        now = mexico_now()
        for key, value in _dump(data).items():
            setattr(row, key, value)
        row.nombre = row.nombre.strip()
        row.telefono = row.telefono.strip() if row.telefono else None
        row.email = row.email.strip().lower() if row.email else None
        row.rfc = row.rfc.strip().upper() if row.rfc else None
        row.fecha_actualizacion = now
        session.add(row)
        _audit(
            session,
            action="EDITAR_CLIENTE",
            entity="cliente",
            entity_id=client_id,
            details={"nombre": row.nombre, "activo": row.activo},
            user=_name(current_user),
            now=now,
        )
        session.commit()
        return _serialize_client(row, session)

    @app.get("/pos/v3/clientes/{client_id}/estado-cuenta")
    def client_statement(
        client_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(POSCliente, client_id)
        if not row:
            raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        credits = session.exec(
            select(POSCredito)
            .where(POSCredito.id_cliente == client_id)
            .order_by(POSCredito.fecha_creacion.desc())
        ).all()
        result = []
        for credit in credits:
            sale = session.get(VentaPOS, credit.id_venta)
            installments = session.exec(
                select(POSAbono)
                .where(POSAbono.id_credito == credit.id_credito)
                .order_by(POSAbono.fecha.asc())
            ).all()
            result.append(
                {
                    "id_credito": credit.id_credito,
                    "id_venta": credit.id_venta,
                    "folio_venta": sale.folio if sale else None,
                    "total_credito": _money(credit.total_credito),
                    "saldo": _money(credit.saldo),
                    "fecha_vencimiento": credit.fecha_vencimiento,
                    "estado": credit.estado,
                    "fecha_creacion": credit.fecha_creacion,
                    "abonos": [
                        {
                            "id_abono": item.id_abono,
                            "folio": item.folio,
                            "metodo": item.metodo,
                            "monto": _money(item.monto),
                            "referencia": item.referencia,
                            "usuario": item.usuario,
                            "fecha": item.fecha,
                        }
                        for item in installments
                    ],
                }
            )
        return {"cliente": _serialize_client(row, session), "creditos": result}

    # -------------------- CONFIGURACIÓN / PRECIOS --------------------

    @app.get("/pos/v3/productos/unidad/{sku}")
    def get_product_unit(
        sku: str,
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        clean_sku = sku.strip()
        if not clean_sku:
            raise HTTPException(status_code=400, detail="SKU obligatorio.")
        product = session.exec(
            select(Producto).where(Producto.sku == clean_sku)
        ).first()
        row = _product_config(session, clean_sku)
        return _serialize_product_unit(
            clean_sku,
            row,
            product_exists=product is not None,
        )

    @app.put("/pos/v3/productos/unidad/{sku}")
    def save_product_unit(
        sku: str,
        data: ProductoUnidadRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        clean_sku = sku.strip()
        if not clean_sku:
            raise HTTPException(status_code=400, detail="SKU obligatorio.")

        unit_key, definition = _product_unit(data.unidad_venta)
        product = session.exec(
            select(Producto).where(Producto.sku == clean_sku)
        ).first()
        existing = _product_config(session, clean_sku)

        if existing:
            existing_key, _ = _product_unit(existing.unidad_venta)
            if (
                product is not None
                and int(product.cantidad or 0) > 0
                and existing_key != unit_key
            ):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "No se puede cambiar la unidad de un producto con existencias. "
                        "Deja el stock en cero o utiliza un SKU nuevo."
                    ),
                )

        now = mexico_now()
        row = _product_config(
            session,
            clean_sku,
            create=True,
            now=now,
            user=_name(current_user),
        )
        assert row is not None

        row.unidad_venta = definition["unidad_venta"]
        row.permite_fraccion = definition["permite_fraccion"]
        row.factor_inventario = definition["factor_inventario"]
        row.activo = True
        row.fecha_actualizacion = now
        row.actualizado_por = _name(current_user)
        session.add(row)

        _audit(
            session,
            action="CONFIGURAR_UNIDAD_PRODUCTO",
            entity="producto",
            entity_id=clean_sku,
            details={
                "unidad_venta": definition["unidad_venta"],
                "unidad_interna": definition["unidad_interna"],
                "factor_inventario": definition["factor_inventario"],
            },
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)

        return _serialize_product_unit(
            clean_sku,
            row,
            product_exists=product is not None,
        )

    @app.get("/pos/v3/productos/configuracion")
    def list_product_configs(
        query: str = Query(default="", max_length=100),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        statement = select(POSProductoConfiguracion).order_by(
            POSProductoConfiguracion.sku.asc()
        )
        if query.strip():
            statement = statement.where(
                POSProductoConfiguracion.sku.ilike(f"%{query.strip()}%")
            )
        return session.exec(statement).all()

    @app.put("/pos/v3/productos/configuracion/{sku}")
    def save_product_config(
        sku: str,
        data: ProductoConfigRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        product = session.exec(select(Producto).where(Producto.sku == sku)).first()
        if not product:
            raise HTTPException(status_code=404, detail="Producto no encontrado.")
        if data.sku.strip() != sku:
            raise HTTPException(status_code=400, detail="El SKU del cuerpo no coincide.")
        if data.permite_fraccion and data.factor_inventario <= 1:
            raise HTTPException(
                status_code=400,
                detail="Un producto fraccionado debe usar factor_inventario mayor a 1.",
            )
        now = mexico_now()
        row = _product_config(session, sku, create=True, now=now, user=_name(current_user))
        assert row is not None
        for key, value in _dump(data).items():
            setattr(row, key, value)
        row.sku = sku
        row.unidad_venta = row.unidad_venta.strip().lower()
        row.fecha_actualizacion = now
        row.actualizado_por = _name(current_user)
        session.add(row)
        _audit(
            session,
            action="CONFIGURAR_PRODUCTO_POS",
            entity="producto",
            entity_id=sku,
            details=_dump(data),
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)
        return row

    @app.put("/pos/v3/precios-cliente")
    def save_customer_price(
        data: PrecioClienteRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        _client(session, data.id_cliente)
        if not session.exec(select(Producto).where(Producto.sku == data.sku)).first():
            raise HTTPException(status_code=404, detail="Producto no encontrado.")
        row = session.exec(
            select(POSPrecioCliente).where(
                (POSPrecioCliente.id_cliente == data.id_cliente)
                & (POSPrecioCliente.sku == data.sku)
            )
        ).first()
        now = mexico_now()
        if row is None:
            row = POSPrecioCliente(
                id_cliente=data.id_cliente,
                sku=data.sku,
                precio=data.precio,
                activo=data.activo,
                fecha_actualizacion=now,
                actualizado_por=_name(current_user),
            )
        else:
            row.precio = data.precio
            row.activo = data.activo
            row.fecha_actualizacion = now
            row.actualizado_por = _name(current_user)
        session.add(row)
        session.commit()
        session.refresh(row)
        return row

    # -------------------- PROMOCIONES --------------------

    @app.get("/pos/v3/promociones")
    def list_promotions(
        include_inactive: bool = Query(default=True),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        statement = select(POSPromocion).order_by(
            POSPromocion.prioridad.desc(), POSPromocion.fecha_creacion.desc()
        )
        if not include_inactive:
            statement = statement.where(POSPromocion.activa == True)  # noqa: E712
        return session.exec(statement).all()

    @app.post("/pos/v3/promociones")
    def create_promotion(
        data: PromocionRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        promo_type = data.tipo.strip().upper()
        if promo_type not in {"PORCENTAJE", "PRECIO_FIJO", "NXM"}:
            raise HTTPException(status_code=400, detail="Tipo de promoción inválido.")
        if promo_type == "PORCENTAJE" and data.porcentaje <= 0:
            raise HTTPException(status_code=400, detail="Indica el porcentaje.")
        if promo_type == "PRECIO_FIJO" and data.precio_fijo <= 0:
            raise HTTPException(status_code=400, detail="Indica el precio fijo.")
        if promo_type == "NXM" and not (
            data.compra_cantidad > 0
            and data.paga_cantidad >= 0
            and data.paga_cantidad < data.compra_cantidad
        ):
            raise HTTPException(status_code=400, detail="Configura correctamente N por M.")
        now = mexico_now()
        payload = _dump(data)
        payload["tipo"] = promo_type
        payload["sku"] = data.sku.strip() if data.sku else None
        row = POSPromocion(
            **payload,
            fecha_creacion=now,
            creada_por=_name(current_user),
        )
        session.add(row)
        session.flush()
        _audit(
            session,
            action="CREAR_PROMOCION",
            entity="promocion",
            entity_id=row.id_promocion,
            details=_dump(data),
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)
        return row

    @app.put("/pos/v3/promociones/{promotion_id}")
    def update_promotion(
        promotion_id: int,
        data: PromocionRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(POSPromocion, promotion_id)
        if not row:
            raise HTTPException(status_code=404, detail="Promoción no encontrada.")
        for key, value in _dump(data).items():
            setattr(row, key, value)
        row.tipo = row.tipo.strip().upper()
        row.sku = row.sku.strip() if row.sku else None
        session.add(row)
        _audit(
            session,
            action="EDITAR_PROMOCION",
            entity="promocion",
            entity_id=promotion_id,
            details=_dump(data),
            user=_name(current_user),
            now=mexico_now(),
        )
        session.commit()
        return row

    # -------------------- BÚSQUEDA / COTIZACIÓN --------------------

    @app.get("/pos/v3/productos/buscar")
    def search_products(
        query: str = Query(..., min_length=1),
        id_cliente: Optional[int] = Query(default=None),
        limite: int = Query(default=20, ge=1, le=50),
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        value = query.strip()
        exact = session.exec(
            select(Producto).where(
                or_(Producto.sku == value, Producto.codigo_barras == value)
            )
        ).first()
        if exact:
            products = [exact]
        else:
            pattern = f"%{value}%"
            products = session.exec(
                select(Producto)
                .where(
                    or_(
                        Producto.sku.ilike(pattern),
                        Producto.nombre.ilike(pattern),
                        Producto.codigo_barras.ilike(pattern),
                    )
                )
                .limit(limite)
            ).all()
        skus = [product.sku for product in products]
        configs_by_sku: Dict[str, POSProductoConfiguracion] = {}
        specials_by_sku: Dict[str, POSPrecioCliente] = {}

        if skus:
            configs = session.exec(
                select(POSProductoConfiguracion).where(
                    POSProductoConfiguracion.sku.in_(skus)
                )
            ).all()
            configs_by_sku = {row.sku: row for row in configs}

            if id_cliente is not None:
                special_rows = session.exec(
                    select(POSPrecioCliente).where(
                        (POSPrecioCliente.id_cliente == id_cliente)
                        & (POSPrecioCliente.sku.in_(skus))
                        & (POSPrecioCliente.activo == True)  # noqa: E712
                    )
                ).all()
                specials_by_sku = {row.sku: row for row in special_rows}

        result = []
        for product in products:
            config = configs_by_sku.get(product.sku)
            factor = float(
                config.factor_inventario
                if config and config.activo
                else 1
            ) or 1
            special = specials_by_sku.get(product.sku)
            price = (
                special.precio
                if special
                else config.precio_normal
                if config and config.precio_normal is not None
                else product.precio_venta_sugerido
            )
            result.append(
                {
                    "id_producto": product.id_producto,
                    "sku": product.sku,
                    "codigo_barras": product.codigo_barras,
                    "nombre": product.nombre,
                    "descripcion": product.descripcion,
                    "cantidad": int(product.cantidad or 0),
                    "cantidad_disponible_venta": _qty((product.cantidad or 0) / factor),
                    "precio_venta_sugerido": _money(price),
                    "costo_proveedor": _money(product.costo_proveedor),
                    "ubicacion": f"{product.rack}-{product.nivel}-{product.slot}",
                    "rack": product.rack,
                    "nivel": product.nivel,
                    "slot": product.slot,
                    "caducidad": product.caducidad,
                    "unidad_venta": config.unidad_venta if config else "pieza",
                    "permite_fraccion": bool(config.permite_fraccion) if config else False,
                    "factor_inventario": factor,
                    "precio_mayoreo": _money(config.precio_mayoreo) if config and config.precio_mayoreo is not None else None,
                    "cantidad_mayoreo": _qty(config.cantidad_mayoreo) if config else 0,
                    "precio_minimo": _money(config.precio_minimo) if config and config.precio_minimo is not None else None,
                }
            )
        return result

    @app.post("/pos/v3/cotizar")
    def quote_sale(
        data: VentaV3Request,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        client = _client(session, data.id_cliente)
        now = mexico_now()
        is_admin = _role(current_user) == "admin"
        lines = []
        total = 0.0
        for item in data.items:
            product = session.exec(select(Producto).where(Producto.sku == item.sku)).first()
            if not product:
                raise HTTPException(status_code=404, detail=f"Producto {item.sku} no encontrado.")
            config = _product_config(session, product.sku)
            pricing = _calculate_price(
                session,
                product=product,
                config=config,
                client_id=client.id_cliente if client else None,
                quantity=item.cantidad,
                manual_discount=item.descuento_porcentaje,
                now=now,
                is_admin=is_admin,
            )
            total += pricing["final_total"]
            lines.append({"sku": item.sku, "nombre": product.nombre, **pricing})
        return {"items": lines, "total": _money(total)}

    # -------------------- VENTA V3 --------------------

    @app.post("/pos/v3/ventas")
    def create_sale_v3(
        data: VentaV3Request,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        cash_session = _require_open_session(session, current_user)
        existing = session.exec(
            select(POSVentaControl).where(POSVentaControl.operacion_id == data.operacion_id)
        ).first()
        if existing:
            sale = session.get(VentaPOS, existing.id_venta)
            if sale:
                response = _serialize_sale(session, sale, True)
                response["duplicada"] = True
                response["mensaje"] = "La venta ya había sido registrada."
                return response

        sale_type = data.tipo_venta.strip().upper()
        if sale_type not in {"CONTADO", "CREDITO", "PARCIAL"}:
            raise HTTPException(status_code=400, detail="Tipo de venta inválido.")
        client = _client(session, data.id_cliente)
        if sale_type in {"CREDITO", "PARCIAL"} and client is None:
            raise HTTPException(status_code=400, detail="Selecciona un cliente para la venta a crédito.")

        now = mexico_now()
        is_admin = _role(current_user) == "admin"
        max_discount = 100.0 if is_admin else 10.0
        grouped: Dict[str, VentaItemV3Request] = {}
        for item in data.items:
            sku = item.sku.strip()
            if item.descuento_porcentaje > max_discount:
                raise HTTPException(
                    status_code=403,
                    detail=f"Tu rol permite máximo {max_discount:.0f}% de descuento manual.",
                )
            current = grouped.get(sku)
            if current:
                if abs(current.descuento_porcentaje - item.descuento_porcentaje) > 0.001:
                    raise HTTPException(status_code=400, detail=f"El SKU {sku} tiene descuentos distintos.")
                current.cantidad = _qty(current.cantidad + item.cantidad)
            else:
                grouped[sku] = VentaItemV3Request(**{**_dump(item), "sku": sku})

        calculated: List[Dict[str, Any]] = []
        list_total = discount_total = auto_discount_total = total = cost_total = 0.0
        try:
            for item in grouped.values():
                statement = select(Producto).where(Producto.sku == item.sku)
                try:
                    statement = statement.with_for_update()
                except AttributeError:
                    pass
                product = session.exec(statement).first()
                if not product:
                    raise HTTPException(status_code=404, detail=f"Producto {item.sku} no encontrado.")
                config = _product_config(session, product.sku)
                factor = float(config.factor_inventario if config and config.activo else 1) or 1
                if not (config and config.permite_fraccion) and abs(item.cantidad - round(item.cantidad)) > 0.000001:
                    raise HTTPException(status_code=400, detail=f"{product.nombre} solo admite cantidades enteras.")
                inventory_exact = item.cantidad * factor
                inventory_amount = int(round(inventory_exact))
                if abs(inventory_exact - inventory_amount) > 0.000001:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"La cantidad de {product.nombre} no coincide con su factor de inventario. "
                            f"Usa múltiplos de {1 / factor:g} {config.unidad_venta if config else 'pieza'}."
                        ),
                    )
                if inventory_amount <= 0 or inventory_amount > int(product.cantidad or 0):
                    raise HTTPException(
                        status_code=400,
                        detail=f"Stock insuficiente para {product.nombre}. Disponible: {(product.cantidad or 0) / factor:g}.",
                    )
                pricing = _calculate_price(
                    session,
                    product=product,
                    config=config,
                    client_id=client.id_cliente if client else None,
                    quantity=item.cantidad,
                    manual_discount=item.descuento_porcentaje,
                    now=now,
                    is_admin=is_admin,
                )
                list_total += pricing["line_list"]
                auto_discount_total += pricing["automatic_discount"]
                discount_total += pricing["automatic_discount"] + pricing["manual_discount_amount"]
                total += pricing["final_total"]
                cost_total += pricing["cost_total"]
                calculated.append(
                    {
                        "request": item,
                        "product": product,
                        "config": config,
                        "inventory_amount": inventory_amount,
                        "pricing": pricing,
                    }
                )

            list_total = _money(list_total)
            auto_discount_total = _money(auto_discount_total)
            discount_total = _money(discount_total)
            total = _money(total)
            cost_total = _money(cost_total)

            valid_methods = {"efectivo", "tarjeta", "transferencia"}
            payments = []
            paid = cash_applied = 0.0
            for payment in data.pagos:
                method = payment.metodo.strip().lower()
                if method not in valid_methods:
                    raise HTTPException(status_code=400, detail="Método de pago inválido.")
                amount = _money(payment.monto)
                paid = _money(paid + amount)
                if method == "efectivo":
                    cash_applied = _money(cash_applied + amount)
                payments.append(
                    {
                        "metodo": method,
                        "monto": amount,
                        "referencia": payment.referencia.strip() if payment.referencia else None,
                    }
                )

            if sale_type == "CONTADO" and abs(paid - total) > 0.01:
                raise HTTPException(status_code=400, detail=f"Los pagos deben sumar ${total:.2f}.")
            if sale_type == "CREDITO" and paid >= total - 0.01:
                raise HTTPException(status_code=400, detail="Una venta a crédito debe conservar saldo pendiente.")
            if sale_type == "PARCIAL" and not (0 < paid < total - 0.01):
                raise HTTPException(status_code=400, detail="El pago parcial debe ser mayor a cero y menor al total.")
            if paid > total + 0.01:
                raise HTTPException(status_code=400, detail="El pago aplicado supera el total de la venta.")

            pending = _money(total - paid)
            if pending > 0 and client:
                open_credits = session.exec(
                    select(POSCredito).where(
                        (POSCredito.id_cliente == client.id_cliente)
                        & (POSCredito.estado != "PAGADO")
                        & (POSCredito.estado != "CANCELADO")
                    )
                ).all()
                current_balance = _money(sum(row.saldo for row in open_credits))
                if not is_admin and client.limite_credito > 0 and current_balance + pending > client.limite_credito:
                    raise HTTPException(
                        status_code=403,
                        detail=(
                            f"El crédito excede el límite del cliente. Disponible: "
                            f"${max(client.limite_credito - current_balance, 0):.2f}."
                        ),
                    )

            received = _money(cash_applied if data.efectivo_recibido is None else data.efectivo_recibido)
            if cash_applied > 0 and received < cash_applied:
                raise HTTPException(status_code=400, detail="El efectivo recibido no cubre el pago en efectivo.")
            if cash_applied <= 0:
                received = 0
            change = _money(max(received - cash_applied, 0))

            sale = VentaPOS(
                folio=_folio("RN", now),
                usuario=_name(current_user),
                subtotal=list_total,
                descuento_total=discount_total,
                total=total,
                costo_total=cost_total,
                ganancia=_money(total - cost_total),
                efectivo_recibido=received,
                cambio=change,
                estado="COMPLETADA",
                fecha=now,
            )
            session.add(sale)
            session.flush()
            if sale.id_venta is None:
                raise RuntimeError("No fue posible generar el identificador de venta.")

            extra = POSVentaExtra(
                id_venta=sale.id_venta,
                id_cliente=client.id_cliente if client else None,
                cliente_nombre=client.nombre if client else None,
                tipo_venta=sale_type,
                saldo_pendiente=pending,
                fecha_vencimiento=(
                    data.fecha_vencimiento
                    if pending > 0 and data.fecha_vencimiento
                    else now.date() + timedelta(days=client.dias_credito if client else 0)
                    if pending > 0
                    else None
                ),
                descuento_promociones=auto_discount_total,
                promociones_json="[]",
                fecha_creacion=now,
            )
            session.add(extra)
            promotions_used = []

            for row in calculated:
                item = row["request"]
                product = row["product"]
                config = row["config"]
                pricing = row["pricing"]
                inventory_amount = row["inventory_amount"]
                lots = descontar_lotes_fefo(
                    session=session,
                    sku=product.sku,
                    cantidad=inventory_amount,
                )
                product.cantidad = int(product.cantidad or 0) - inventory_amount
                product.caducidad = obtener_caducidad_mas_proxima(session, product.sku)
                product.ultima_actualizacion = now
                session.add(product)

                detail_row = VentaPOSDetalle(
                    id_venta=sale.id_venta,
                    id_producto=product.id_producto,
                    sku=product.sku,
                    codigo_barras=product.codigo_barras,
                    nombre=product.nombre,
                    cantidad=inventory_amount,
                    precio_lista=pricing["base_unit"],
                    descuento_porcentaje=item.descuento_porcentaje,
                    precio_unitario_final=pricing["final_unit"],
                    subtotal=pricing["final_total"],
                    costo_unitario=pricing["cost_unit"],
                    costo_total=pricing["cost_total"],
                    ganancia=pricing["profit"],
                )
                session.add(detail_row)
                session.flush()
                if detail_row.id_detalle is None:
                    raise RuntimeError("No se generó el detalle de venta.")
                session.add(
                    POSVentaDetalleExtra(
                        id_detalle=detail_row.id_detalle,
                        id_venta=sale.id_venta,
                        unidad_venta=config.unidad_venta if config else "pieza",
                        cantidad_venta=item.cantidad,
                        cantidad_inventario=inventory_amount,
                        factor_inventario=pricing["factor"],
                        descuento_automatico=pricing["automatic_discount"],
                        promocion_nombre=pricing["promotion_name"],
                        precio_origen=pricing["origin"],
                    )
                )
                if pricing["promotion_name"]:
                    promotions_used.append(
                        {
                            "sku": product.sku,
                            "promocion": pricing["promotion_name"],
                            "descuento": pricing["automatic_discount"],
                        }
                    )

                for lot in lots:
                    session.add(
                        POSVentaLote(
                            id_venta=sale.id_venta,
                            id_detalle=detail_row.id_detalle,
                            id_lote=lot.get("id_lote"),
                            sku=product.sku,
                            cantidad=int(lot.get("cantidad_descontada") or 0),
                            cantidad_restaurada=0,
                        )
                    )

                movement = Movimiento(
                    accion="Egreso",
                    sku=product.sku,
                    producto=product.nombre,
                    cantidad=inventory_amount,
                    ubicacion=f"{product.rack}-{product.nivel}-{product.slot}",
                    usuario=_name(current_user),
                    fecha=now,
                    costo_proveedor=pricing["cost_unit"],
                    precio_venta=pricing["final_unit"],
                    ingreso_total=pricing["final_total"],
                    costo_total=pricing["cost_total"],
                    ganancia=pricing["profit"],
                )
                session.add(movement)
                session.flush()
                if movement.id_mov is not None:
                    session.add(
                        POSVentaMovimiento(
                            id_venta=sale.id_venta,
                            id_detalle=detail_row.id_detalle,
                            id_movimiento=movement.id_mov,
                        )
                    )

            extra.promociones_json = json.dumps(promotions_used, ensure_ascii=False)
            session.add(extra)
            for payment in payments:
                session.add(VentaPOSPago(id_venta=sale.id_venta, **payment))
            control = POSVentaControl(
                id_venta=sale.id_venta,
                id_sesion=int(cash_session.id_sesion or 0),
                operacion_id=data.operacion_id,
                fecha_creacion=now,
            )
            session.add(control)

            if pending > 0 and client:
                due = extra.fecha_vencimiento or now.date()
                session.add(
                    POSCredito(
                        id_venta=sale.id_venta,
                        id_cliente=int(client.id_cliente or 0),
                        total_credito=pending,
                        saldo=pending,
                        fecha_vencimiento=due,
                        estado="PENDIENTE",
                        usuario_autorizo=_name(current_user),
                        fecha_creacion=now,
                        fecha_actualizacion=now,
                    )
                )

            _audit(
                session,
                action="VENTA_POS_V3",
                entity="venta",
                entity_id=sale.id_venta,
                details={
                    "folio": sale.folio,
                    "total": total,
                    "tipo_venta": sale_type,
                    "cliente": client.nombre if client else None,
                    "saldo": pending,
                },
                user=_name(current_user),
                now=now,
            )
            session.commit()
            session.refresh(sale)
            response = _serialize_sale(session, sale, True)
            response["mensaje"] = "Venta registrada correctamente."
            response["duplicada"] = False
            return response
        except HTTPException:
            session.rollback()
            raise
        except Exception as exc:
            session.rollback()
            print(f"❌ Error venta POS Fase 3: {exc}")
            raise HTTPException(status_code=500, detail=f"No se pudo registrar la venta: {exc}") from exc

    @app.get("/pos/v3/ventas")
    def list_sales_v3(
        limite: int = Query(default=100, ge=1, le=500),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        rows = session.exec(
            select(VentaPOS)
            .order_by(VentaPOS.fecha.desc())
            .limit(limite)
        ).all()

        sale_ids = [
            int(row.id_venta)
            for row in rows
            if row.id_venta is not None
        ]
        extras_by_sale: Dict[int, POSVentaExtra] = {}
        controls_by_sale: Dict[int, POSVentaControl] = {}

        if sale_ids:
            extras = session.exec(
                select(POSVentaExtra).where(
                    POSVentaExtra.id_venta.in_(sale_ids)
                )
            ).all()
            extras_by_sale = {row.id_venta: row for row in extras}

            controls = session.exec(
                select(POSVentaControl).where(
                    POSVentaControl.id_venta.in_(sale_ids)
                )
            ).all()
            controls_by_sale = {row.id_venta: row for row in controls}

        result = []
        for sale in rows:
            sale_id = int(sale.id_venta or 0)
            extra = extras_by_sale.get(sale_id)
            control = controls_by_sale.get(sale_id)
            result.append(
                {
                    "id_venta": sale.id_venta,
                    "folio": sale.folio,
                    "usuario": sale.usuario,
                    "subtotal": _money(sale.subtotal),
                    "descuento_total": _money(sale.descuento_total),
                    "total": _money(sale.total),
                    "costo_total": _money(sale.costo_total),
                    "ganancia": _money(sale.ganancia),
                    "efectivo_recibido": _money(sale.efectivo_recibido),
                    "cambio": _money(sale.cambio),
                    "estado": sale.estado,
                    "fecha": sale.fecha,
                    "id_cliente": extra.id_cliente if extra else None,
                    "cliente_nombre": extra.cliente_nombre if extra else None,
                    "tipo_venta": extra.tipo_venta if extra else "CONTADO",
                    "saldo_pendiente": _money(
                        extra.saldo_pendiente if extra else 0
                    ),
                    "fecha_vencimiento": (
                        extra.fecha_vencimiento if extra else None
                    ),
                    "descuento_promociones": _money(
                        extra.descuento_promociones if extra else 0
                    ),
                    "id_sesion": control.id_sesion if control else None,
                    "motivo_anulacion": (
                        control.motivo_anulacion if control else None
                    ),
                    "fecha_anulacion": (
                        control.fecha_anulacion if control else None
                    ),
                }
            )

        return result

    @app.get("/pos/v3/ventas/{sale_id}")
    def get_sale_v3(
        sale_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(VentaPOS, sale_id)
        if not row:
            raise HTTPException(status_code=404, detail="Venta no encontrada.")
        return _serialize_sale(session, row, True)

    # -------------------- ABONOS --------------------

    @app.get("/pos/v3/creditos")
    def list_credits(
        estado: Optional[str] = Query(default=None),
        client_id: Optional[int] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        statement = select(POSCredito).order_by(
            POSCredito.fecha_vencimiento.asc()
        )
        if estado:
            statement = statement.where(POSCredito.estado == estado.upper())
        if client_id:
            statement = statement.where(POSCredito.id_cliente == client_id)

        rows = session.exec(statement).all()
        client_ids = list({row.id_cliente for row in rows})
        sale_ids = list({row.id_venta for row in rows})
        clients_by_id: Dict[int, POSCliente] = {}
        sales_by_id: Dict[int, VentaPOS] = {}

        if client_ids:
            clients = session.exec(
                select(POSCliente).where(
                    POSCliente.id_cliente.in_(client_ids)
                )
            ).all()
            clients_by_id = {
                int(row.id_cliente or 0): row
                for row in clients
            }

        if sale_ids:
            sales = session.exec(
                select(VentaPOS).where(
                    VentaPOS.id_venta.in_(sale_ids)
                )
            ).all()
            sales_by_id = {
                int(row.id_venta or 0): row
                for row in sales
            }

        result = []
        for row in rows:
            client = clients_by_id.get(row.id_cliente)
            sale = sales_by_id.get(row.id_venta)
            result.append(
                {
                    **_dump(row),
                    "saldo": _money(row.saldo),
                    "total_credito": _money(row.total_credito),
                    "cliente_nombre": client.nombre if client else None,
                    "folio_venta": sale.folio if sale else None,
                }
            )

        return result

    @app.post("/pos/v3/creditos/{credit_id}/abonos")
    def add_installment(
        credit_id: int,
        data: AbonoRequest,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        credit = session.get(POSCredito, credit_id)
        if not credit or credit.estado in {"PAGADO", "CANCELADO"}:
            raise HTTPException(status_code=404, detail="Crédito no disponible.")
        amount = _money(data.monto)
        if amount > credit.saldo + 0.01:
            raise HTTPException(status_code=400, detail=f"El saldo es ${credit.saldo:.2f}.")
        method = data.metodo.strip().lower()
        if method not in {"efectivo", "tarjeta", "transferencia"}:
            raise HTTPException(status_code=400, detail="Método inválido.")
        cash_session = _require_open_session(session, current_user)
        now = mexico_now()
        row = POSAbono(
            folio=_folio("AB", now),
            id_credito=int(credit.id_credito or 0),
            id_cliente=credit.id_cliente,
            id_sesion=cash_session.id_sesion,
            metodo=method,
            monto=amount,
            referencia=data.referencia.strip() if data.referencia else None,
            usuario=_name(current_user),
            fecha=now,
        )
        session.add(row)
        credit.saldo = _money(credit.saldo - amount)
        credit.estado = "PAGADO" if credit.saldo <= 0.01 else "PARCIAL"
        credit.fecha_actualizacion = now
        session.add(credit)
        extra = _sale_extra(session, credit.id_venta)
        if extra:
            extra.saldo_pendiente = credit.saldo
            session.add(extra)
        _audit(
            session,
            action="ABONO_CREDITO",
            entity="credito",
            entity_id=credit_id,
            details={"monto": amount, "metodo": method, "saldo": credit.saldo},
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)
        return {
            "mensaje": "Abono registrado.",
            "abono": row,
            "credito": credit,
            "sesion": _cash_summary(session, cash_session),
        }

    # -------------------- CAJA V3 --------------------

    @app.get("/pos/v3/caja/sesion-actual")
    def current_cash_v3(
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = _open_session(session, current_user)
        return {"abierta": row is not None, "sesion": _cash_summary(session, row) if row else None}

    @app.get(
        "/pos/v3/caja/sesiones/{session_id}/resumen"
    )
    def session_report_v3(
        session_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(POSSesionCaja, session_id)

        if not row:
            raise HTTPException(
                status_code=404,
                detail="Sesión de caja no encontrada.",
            )

        is_admin = _role(current_user) == "admin"
        current_keys = {
            _key(current_user),
            _name(current_user),
        }

        if not is_admin and row.usuario not in current_keys:
            raise HTTPException(
                status_code=403,
                detail=(
                    "Solo puedes consultar el resumen "
                    "de tu propia caja."
                ),
            )

        return _session_report(
            session,
            row,
            Movimiento,
        )

    @app.post("/pos/v3/caja/cerrar")
    def close_cash_v3(
        data: CerrarCajaV3Request,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = _require_open_session(session, current_user)
        summary = _cash_summary(session, row)
        now = mexico_now()
        row.efectivo_esperado = summary["efectivo_esperado"]
        row.efectivo_contado = _money(data.efectivo_contado)
        row.diferencia = _money(row.efectivo_contado - row.efectivo_esperado)
        row.fecha_cierre = now
        row.estado = "CERRADA"
        row.observaciones = data.observaciones.strip() if data.observaciones else None
        session.add(row)
        _audit(
            session,
            action="CERRAR_CAJA",
            entity="sesion_caja",
            entity_id=row.id_sesion,
            details={
                "esperado": row.efectivo_esperado,
                "contado": row.efectivo_contado,
                "diferencia": row.diferencia,
            },
            user=_name(current_user),
            now=now,
        )
        session.commit()
        session.refresh(row)

        return {
            "mensaje": "Caja cerrada correctamente.",
            "sesion": _cash_summary(session, row),
        }

    # -------------------- CANCELACIONES / DEVOLUCIONES V3 --------------------

    @app.post("/pos/v3/ventas/{sale_id}/cancelar")
    def cancel_sale_v3(
        sale_id: int,
        data: CancelarV3Request,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        sale = session.get(VentaPOS, sale_id)
        if not sale:
            raise HTTPException(status_code=404, detail="Venta no encontrada.")
        if sale.estado == "CANCELADA":
            return _serialize_sale(session, sale, True)
        credit = _credit_by_sale(session, sale_id)
        if credit:
            installments = session.exec(
                select(POSAbono).where(POSAbono.id_credito == credit.id_credito)
            ).all()
            if installments:
                raise HTTPException(
                    status_code=409,
                    detail="La venta ya tiene abonos. Usa devoluciones para conservar la auditoría.",
                )
        now = mexico_now()
        details = session.exec(
            select(VentaPOSDetalle).where(VentaPOSDetalle.id_venta == sale_id)
        ).all()
        try:
            for detail in details:
                extra = _detail_extra(session, int(detail.id_detalle or 0))
                inventory = int(extra.cantidad_inventario if extra else detail.cantidad or 0)
                already = _returned_inventory(session, int(detail.id_detalle or 0))
                remaining = max(inventory - already, 0)
                if remaining <= 0:
                    continue
                product = session.exec(select(Producto).where(Producto.sku == detail.sku)).first()
                if product:
                    product.cantidad = int(product.cantidad or 0) + remaining
                    product.caducidad = obtener_caducidad_mas_proxima(session, product.sku)
                    product.ultima_actualizacion = now
                    session.add(product)
                _restore_lots(session, ProductoLote, sale_id, int(detail.id_detalle or 0), remaining)
                _update_movement_net(session, Movimiento, detail, inventory)
            sale.estado = "CANCELADA"
            session.add(sale)
            control = _sale_control(session, sale_id)
            if control:
                control.anulada_por = _name(current_user)
                control.motivo_anulacion = data.motivo.strip()
                control.fecha_anulacion = now
                session.add(control)
            if credit:
                credit.saldo = 0
                credit.estado = "CANCELADO"
                credit.fecha_actualizacion = now
                session.add(credit)
                extra_sale = _sale_extra(session, sale_id)
                if extra_sale:
                    extra_sale.saldo_pendiente = 0
                    session.add(extra_sale)
            _audit(
                session,
                action="CANCELAR_VENTA",
                entity="venta",
                entity_id=sale_id,
                details={"motivo": data.motivo},
                user=_name(current_user),
                now=now,
            )
            session.commit()
            return _serialize_sale(session, sale, True)
        except HTTPException:
            session.rollback()
            raise
        except Exception as exc:
            session.rollback()
            raise HTTPException(status_code=500, detail=f"No se pudo cancelar: {exc}") from exc

    @app.post("/pos/v3/ventas/{sale_id}/devoluciones")
    def return_sale_v3(
        sale_id: int,
        data: DevolucionV3Request,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        sale = session.get(VentaPOS, sale_id)
        if not sale or sale.estado == "CANCELADA":
            raise HTTPException(status_code=404, detail="Venta no disponible.")
        cash_session = _require_open_session(session, current_user)
        method = data.metodo_reembolso.strip().lower()
        if method not in {"efectivo", "tarjeta", "transferencia"}:
            raise HTTPException(status_code=400, detail="Método de reembolso inválido.")
        now = mexico_now()
        selected: List[Dict[str, Any]] = []
        total_return = 0.0
        cost_return = 0.0
        for requested in data.items:
            detail = session.get(VentaPOSDetalle, requested.id_detalle)
            if not detail or detail.id_venta != sale_id:
                raise HTTPException(status_code=404, detail="Detalle de venta inválido.")
            extra = _detail_extra(session, int(detail.id_detalle or 0))
            factor = float(extra.factor_inventario if extra else 1) or 1
            inventory_exact = requested.cantidad * factor
            inventory_amount = int(round(inventory_exact))
            if abs(inventory_exact - inventory_amount) > 0.000001:
                raise HTTPException(status_code=400, detail=f"Cantidad inválida para {detail.nombre}.")
            sold_inventory = int(extra.cantidad_inventario if extra else detail.cantidad or 0)
            returned = _returned_inventory(session, int(detail.id_detalle or 0))
            if inventory_amount <= 0 or returned + inventory_amount > sold_inventory:
                raise HTTPException(status_code=400, detail=f"La devolución excede lo vendido de {detail.nombre}.")
            line_total = _money(detail.precio_unitario_final * requested.cantidad)
            line_cost = _money(detail.costo_unitario * requested.cantidad)
            total_return += line_total
            cost_return += line_cost
            selected.append(
                {
                    "detail": detail,
                    "extra": extra,
                    "quantity_sale": requested.cantidad,
                    "inventory_amount": inventory_amount,
                    "line_total": line_total,
                    "line_cost": line_cost,
                    "returned_before": returned,
                }
            )
        total_return = _money(total_return)
        credit = _credit_by_sale(session, sale_id)
        credit_adjustment = 0.0
        real_refund = total_return
        if credit and credit.saldo > 0:
            credit_adjustment = _money(min(total_return, credit.saldo))
            real_refund = _money(total_return - credit_adjustment)

        try:
            return_row = POSDevolucion(
                folio=_folio("DEV", now),
                id_venta=sale_id,
                id_sesion=int(cash_session.id_sesion or 0),
                usuario=_name(current_user),
                motivo=data.motivo.strip(),
                metodo_reembolso=method,
                monto=total_return,
                estado="COMPLETADA",
                fecha=now,
            )
            session.add(return_row)
            session.flush()
            if return_row.id_devolucion is None:
                raise RuntimeError("No se generó la devolución.")
            session.add(
                POSDevolucionExtra(
                    id_devolucion=return_row.id_devolucion,
                    ajuste_credito=credit_adjustment,
                    reembolso_real=real_refund,
                )
            )
            for item in selected:
                detail = item["detail"]
                product = session.exec(select(Producto).where(Producto.sku == detail.sku)).first()
                if product:
                    product.cantidad = int(product.cantidad or 0) + item["inventory_amount"]
                    product.caducidad = obtener_caducidad_mas_proxima(session, product.sku)
                    product.ultima_actualizacion = now
                    session.add(product)
                _restore_lots(
                    session,
                    ProductoLote,
                    sale_id,
                    int(detail.id_detalle or 0),
                    item["inventory_amount"],
                )
                session.add(
                    POSDevolucionDetalle(
                        id_devolucion=return_row.id_devolucion,
                        id_detalle_venta=int(detail.id_detalle or 0),
                        id_producto=detail.id_producto,
                        sku=detail.sku,
                        nombre=detail.nombre,
                        cantidad=item["inventory_amount"],
                        precio_unitario=detail.precio_unitario_final,
                        subtotal=item["line_total"],
                    )
                )
                new_returned = item["returned_before"] + item["inventory_amount"]
                _update_movement_net(session, Movimiento, detail, new_returned)

                movement_location = (
                    f"{product.rack}-{product.nivel}-{product.slot}"
                    if product
                    else "Sin ubicación registrada"
                )

                session.add(
                    Movimiento(
                        accion="Devolución",
                        sku=detail.sku,
                        producto=detail.nombre,
                        cantidad=item["inventory_amount"],
                        ubicacion=movement_location,
                        usuario=_name(current_user),
                        fecha=now,
                        costo_proveedor=_money(
                            detail.costo_unitario
                        ),
                        precio_venta=_money(
                            detail.precio_unitario_final
                        ),
                        ingreso_total=-_money(
                            item["line_total"]
                        ),
                        costo_total=-_money(
                            item["line_cost"]
                        ),
                        ganancia=-_money(
                            item["line_total"]
                            - item["line_cost"]
                        ),
                    )
                )
            if credit and credit_adjustment > 0:
                credit.saldo = _money(credit.saldo - credit_adjustment)
                credit.total_credito = _money(max(credit.total_credito - credit_adjustment, 0))
                credit.estado = "PAGADO" if credit.saldo <= 0.01 else "PARCIAL"
                credit.fecha_actualizacion = now
                session.add(credit)
                extra_sale = _sale_extra(session, sale_id)
                if extra_sale:
                    extra_sale.saldo_pendiente = credit.saldo
                    session.add(extra_sale)
            _audit(
                session,
                action="DEVOLUCION_VENTA",
                entity="venta",
                entity_id=sale_id,
                details={
                    "monto": total_return,
                    "ajuste_credito": credit_adjustment,
                    "reembolso_real": real_refund,
                    "motivo": data.motivo,
                },
                user=_name(current_user),
                now=now,
            )
            session.commit()
            return {
                "mensaje": "Devolución registrada.",
                "id_devolucion": return_row.id_devolucion,
                "monto": total_return,
                "ajuste_credito": credit_adjustment,
                "reembolso_real": real_refund,
                "venta": _serialize_sale(session, sale, True),
            }
        except HTTPException:
            session.rollback()
            raise
        except Exception as exc:
            session.rollback()
            raise HTTPException(status_code=500, detail=f"No se pudo devolver: {exc}") from exc

    # -------------------- REPORTES --------------------

    @app.get("/pos/v3/reportes/diario")
    def daily_report(
        fecha: Optional[date] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        report_date = fecha or mexico_now().date()
        return _daily_report(session, report_date, Movimiento)

    @app.post("/pos/v3/reportes/diario/cerrar")
    def close_daily_report(
        fecha: Optional[date] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        report_date = fecha or mexico_now().date()
        data = _daily_report(session, report_date, Movimiento)
        row = session.exec(
            select(POSReporteDiario).where(POSReporteDiario.fecha_reporte == report_date)
        ).first()
        now = mexico_now()
        if row is None:
            row = POSReporteDiario(
                fecha_reporte=report_date,
                cerrado=True,
                datos_json=json.dumps(data, ensure_ascii=False, default=str),
                generado_por=_name(current_user),
                fecha_generacion=now,
            )
        else:
            row.cerrado = True
            row.datos_json = json.dumps(data, ensure_ascii=False, default=str)
            row.generado_por = _name(current_user)
            row.fecha_generacion = now
        session.add(row)
        _audit(
            session,
            action="CERRAR_REPORTE_DIARIO",
            entity="reporte_diario",
            entity_id=str(report_date),
            details={"ventas_netas": data["resumen"]["ventas_netas"]},
            user=_name(current_user),
            now=now,
        )
        session.commit()
        return {"mensaje": "Reporte diario cerrado.", "reporte": data}

    @app.get("/pos/v3/reportes/diario.xlsx")
    def daily_report_xlsx(
        fecha: Optional[date] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        report_date = fecha or mexico_now().date()
        content = _xlsx_report(_daily_report(session, report_date, Movimiento))
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="racknova-ventas-{report_date}.xlsx"'},
        )

    @app.get("/pos/v3/reportes/diario.pdf")
    def daily_report_pdf(
        fecha: Optional[date] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        report_date = fecha or mexico_now().date()
        content = _pdf_report(_daily_report(session, report_date, Movimiento))
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="racknova-ventas-{report_date}.pdf"'},
        )

    @app.get("/pos/v3/reportes/rango")
    def range_report(
        desde: date,
        hasta: date,
        session: Session = Depends(get_session),
        current_user: Any = Depends(read_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner', 'viewer'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        if hasta < desde or (hasta - desde).days > 366:
            raise HTTPException(status_code=400, detail="Rango inválido o mayor a 366 días.")
        days = []
        current = desde
        while current <= hasta:
            daily = _daily_report(session, current)
            days.append({"fecha": current, **daily["resumen"]})
            current += timedelta(days=1)
        return {
            "desde": desde,
            "hasta": hasta,
            "dias": days,
            "totales": {
                "ventas_netas": _money(sum(row["ventas_netas"] for row in days)),
                "ganancia": _money(sum(row["ganancia"] for row in days)),
                "numero_ventas": sum(row["numero_ventas"] for row in days),
                "devoluciones": sum(row["devoluciones"] for row in days),
            },
        }

    # -------------------- AUDITORÍA --------------------

    @app.get("/pos/v3/auditoria")
    def audit_log(
        limite: int = Query(default=100, ge=1, le=1000),
        accion: Optional[str] = Query(default=None),
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        statement = select(POSAuditoria).order_by(POSAuditoria.fecha.desc()).limit(limite)
        if accion:
            statement = statement.where(POSAuditoria.accion == accion.upper())
        rows = session.exec(statement).all()
        return [
            {
                **_dump(row),
                "detalles": json.loads(row.detalles_json or "{}"),
            }
            for row in rows
        ]

    # ==========================================================
    # RACKNOVA POS V4: CAJAS FIJAS, REPORTES Y MAYOREO
    # ==========================================================

    @app.get("/pos/v4/cajas")
    def fixed_boxes_v4(
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        open_rows = _v4_open_sessions(session)
        box_model = globals().get("POSCaja")
        if box_model is None:
            return []

        boxes = list(session.exec(select(box_model).order_by(box_model.nombre)).all())
        if not boxes:
            _v4_ensure_fixed_box(session, 1)
            _v4_ensure_fixed_box(session, 2)
            boxes = list(session.exec(select(box_model).order_by(box_model.nombre)).all())

        result = []
        for box in boxes:
            if hasattr(box, "activa") and not bool(getattr(box, "activa")):
                continue
            box_id = int(getattr(box, "id_caja", 0) or 0)
            name = str(getattr(box, "nombre", None) or f"Caja {box_id}")
            active = next(
                (
                    row for row in open_rows
                    if int(getattr(row, "id_caja", 0) or 0) == box_id
                    or str(getattr(row, "caja_nombre", "") or "").strip() == name
                ),
                None,
            )
            result.append(
                {
                    "numero": box_id,
                    "id_caja": box_id,
                    "nombre": name,
                    "estado": "ABIERTA" if active else "DISPONIBLE",
                    "sesion": _cash_summary(session, active) if active else None,
                }
            )
        return result

    @app.get("/pos/v4/cajas/abiertas")
    def open_boxes_v4(
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        return [
            _cash_summary(session, row)
            for row in _v4_open_sessions(session)
        ]

    @app.post("/pos/v4/cajas/{box_number}/abrir")
    def open_fixed_box_v4(
        box_number: int,
        payload: Dict[str, Any],
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        box = _v4_ensure_fixed_box(session, box_number)
        box_name = box["nombre"]
        open_rows = _v4_open_sessions(session)

        if any(_v4_session_belongs_to_user(row, current_user) for row in open_rows):
            raise HTTPException(
                status_code=409,
                detail="Tu usuario ya tiene una caja abierta.",
            )

        if any(str(row.caja_nombre or "").strip() == box_name for row in open_rows):
            raise HTTPException(
                status_code=409,
                detail=f"{box_name} ya está siendo utilizada.",
            )

        now = datetime.now()
        initial_fund = max(float(payload.get("fondo_inicial") or 0), 0)
        values = {
            "id_caja": box.get("id_caja"),
            "caja_nombre": box_name,
            "nombre_caja": box_name,
            "usuario": _name(current_user),
            "usuario_id": _key(current_user),
            "fecha_apertura": now,
            "fecha_cierre": None,
            "estado": "ABIERTA",
            "fondo_inicial": initial_fund,
            "total_ventas": 0,
            "efectivo_ventas": 0,
            "efectivo_esperado": initial_fund,
            "efectivo_contado": None,
            "diferencia": None,
            "ventas_completadas": 0,
            "observaciones": payload.get("observaciones"),
        }

        row = POSSesionCaja(**_v4_model_kwargs(POSSesionCaja, values))
        session.add(row)
        session.commit()
        session.refresh(row)

        return {
            "mensaje": f"{box_name} abierta correctamente.",
            "sesion": _cash_summary(session, row),
        }

    @app.get("/pos/v4/caja/sesiones/{session_id}/resumen")
    def session_report_safe_v4(
        session_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        row = session.get(POSSesionCaja, session_id)
        if not row:
            raise HTTPException(status_code=404, detail="Sesión de caja no encontrada.")

        if _role(current_user) != "admin" and not _v4_session_belongs_to_user(row, current_user):
            raise HTTPException(
                status_code=403,
                detail="Solo puedes consultar tu propia caja.",
            )

        return _v4_safe_session_report(session, row, Movimiento)

    @app.get("/pos/v4/reportes/diario")
    def daily_report_v4(
        fecha: str,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        try:
            report_date = date.fromisoformat(fecha)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Fecha inválida. Usa YYYY-MM-DD.") from exc

        start, end = _day_bounds(report_date)
        rows = session.exec(
            select(POSSesionCaja).where(
                POSSesionCaja.fecha_apertura < end,
                (POSSesionCaja.fecha_cierre == None) | (POSSesionCaja.fecha_cierre >= start),
            ).order_by(POSSesionCaja.fecha_apertura.asc())
        ).all()

        reports = [_v4_safe_session_report(session, row, Movimiento) for row in rows]

        def total(key: str) -> float:
            return _money(sum(float(item["totales"].get(key) or 0) for item in reports))

        return {
            "fecha": fecha,
            "totales": {
                "ventas": total("ventas"),
                "devoluciones": total("devoluciones"),
                "ventas_netas": total("ventas_netas"),
                "numero_ventas": sum(int(item["totales"].get("numero_ventas") or 0) for item in reports),
                "numero_devoluciones": sum(int(item["totales"].get("numero_devoluciones") or 0) for item in reports),
                "descuentos": total("descuentos"),
                "costo": total("costo"),
                "ganancia": total("ganancia_antes_devoluciones"),
                "efectivo_esperado": _money(sum(float(item["sesion"].get("efectivo_esperado") or 0) for item in reports)),
                "diferencias": _money(sum(float(item["sesion"].get("diferencia") or 0) for item in reports)),
            },
            "sesiones": reports,
        }

    @app.get("/pos/v4/mayoreo")
    def list_wholesale_v4(
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        _v4_ensure_schema(session)
        rows = session.connection().execute(
            sa_text("SELECT * FROM pos_mayoreo_menudeo WHERE empresa_id = CAST(:empresa AS UUID) ORDER BY nombre, sku")
        , {"empresa": str(session.info.get("racknova_empresa_id") or "11111111-1111-4111-8111-111111111111")}).mappings().all()
        return [_v4_serialize_wholesale(row) for row in rows]

    @app.post("/pos/v4/mayoreo")
    def save_wholesale_v4(
        payload: Dict[str, Any],
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        _v4_ensure_schema(session)
        sku = str(payload.get("sku") or "").strip()
        name = str(payload.get("nombre") or "").strip()
        unit = str(payload.get("unidad") or "").strip().lower()

        if not sku or not name:
            raise HTTPException(status_code=400, detail="SKU y nombre son obligatorios.")
        if unit not in {"kg", "litro"}:
            raise HTTPException(status_code=400, detail="La unidad debe ser kg o litro.")

        retail = float(payload.get("precio_menudeo") or 0)
        wholesale_qty = float(payload.get("cantidad_mayoreo") or 0)
        wholesale_price = float(payload.get("precio_mayoreo") or 0)
        special_qty = payload.get("cantidad_mayoreo_especial")
        special_price = payload.get("precio_mayoreo_especial")

        if retail <= 0 or wholesale_qty <= 0 or wholesale_price <= 0:
            raise HTTPException(status_code=400, detail="Precios y cantidad de mayoreo deben ser mayores a cero.")
        if wholesale_price >= retail:
            raise HTTPException(status_code=400, detail="El precio de mayoreo debe ser menor al menudeo.")
        if (special_qty is None) != (special_price is None):
            raise HTTPException(status_code=400, detail="Completa cantidad y precio del mayoreo especial.")
        if special_qty is not None:
            special_qty = float(special_qty)
            special_price = float(special_price)
            if special_qty <= wholesale_qty:
                raise HTTPException(status_code=400, detail="El mayoreo especial debe iniciar después del mayoreo normal.")
            if special_price >= wholesale_price:
                raise HTTPException(status_code=400, detail="El precio especial debe ser menor al mayoreo normal.")

        params = {
            "empresa": str(session.info.get("racknova_empresa_id") or "11111111-1111-4111-8111-111111111111"),
            "sku": sku,
            "nombre": name,
            "unidad": unit,
            "precio_menudeo": retail,
            "cantidad_mayoreo": wholesale_qty,
            "precio_mayoreo": wholesale_price,
            "cantidad_mayoreo_especial": special_qty,
            "precio_mayoreo_especial": special_price,
            "fecha_inicio": payload.get("fecha_inicio") or None,
            "fecha_fin": payload.get("fecha_fin") or None,
            "activo": bool(payload.get("activo", True)),
        }

        session.connection().execute(
            sa_text(
                """
                INSERT INTO pos_mayoreo_menudeo (
                    empresa_id, sku, nombre, unidad, precio_menudeo,
                    cantidad_mayoreo, precio_mayoreo,
                    cantidad_mayoreo_especial, precio_mayoreo_especial,
                    fecha_inicio, fecha_fin, activo, actualizado_en
                ) VALUES (
                    CAST(:empresa AS UUID), :sku, :nombre, :unidad, :precio_menudeo,
                    :cantidad_mayoreo, :precio_mayoreo,
                    :cantidad_mayoreo_especial, :precio_mayoreo_especial,
                    :fecha_inicio, :fecha_fin, :activo, CURRENT_TIMESTAMP
                )
                ON CONFLICT (empresa_id, sku) DO UPDATE SET
                    nombre = EXCLUDED.nombre,
                    unidad = EXCLUDED.unidad,
                    precio_menudeo = EXCLUDED.precio_menudeo,
                    cantidad_mayoreo = EXCLUDED.cantidad_mayoreo,
                    precio_mayoreo = EXCLUDED.precio_mayoreo,
                    cantidad_mayoreo_especial = EXCLUDED.cantidad_mayoreo_especial,
                    precio_mayoreo_especial = EXCLUDED.precio_mayoreo_especial,
                    fecha_inicio = EXCLUDED.fecha_inicio,
                    fecha_fin = EXCLUDED.fecha_fin,
                    activo = EXCLUDED.activo,
                    actualizado_en = CURRENT_TIMESTAMP
                """
            ),
            params,
        )
        session.commit()
        return {"mensaje": "Regla de mayoreo guardada correctamente."}

    @app.delete("/pos/v4/mayoreo/{rule_id}")
    def delete_wholesale_v4(
        rule_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        _v4_ensure_schema(session)
        result = session.connection().execute(
            sa_text("DELETE FROM pos_mayoreo_menudeo WHERE id_regla = :id AND empresa_id = CAST(:empresa AS UUID)"),
            {"id": rule_id, "empresa": str(session.info.get("racknova_empresa_id") or "11111111-1111-4111-8111-111111111111")},
        )
        session.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Regla no encontrada.")
        return {"mensaje": "Regla de mayoreo eliminada."}

    @app.post("/pos/v4/mayoreo/calcular")
    def calculate_wholesale_v4(
        payload: Dict[str, Any],
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'operator', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        return _v4_wholesale_price(
            session,
            str(payload.get("sku") or "").strip(),
            str(payload.get("unidad") or "").strip(),
            float(payload.get("cantidad") or 0),
            float(payload.get("precio_base") or 0),
        )

    @app.delete("/pos/v4/promociones/{promotion_id}")
    def delete_promotion_v4(
        promotion_id: int,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        promotion_model = globals().get("POSPromocion")
        if promotion_model is None:
            raise HTTPException(status_code=501, detail="Modelo de promociones no disponible.")

        row = session.get(promotion_model, promotion_id)
        if row is None:
            raise HTTPException(status_code=404, detail="Promoción no encontrada.")

        # Elimina primero tablas detalle que tengan id_promocion.
        for model_name, model in list(globals().items()):
            if model is promotion_model or "Promocion" not in model_name:
                continue
            fields = _v4_model_fields(model)
            if "id_promocion" not in fields:
                continue
            try:
                details = session.exec(
                    select(model).where(getattr(model, "id_promocion") == promotion_id)
                ).all()
                for detail in details:
                    session.delete(detail)
            except Exception:
                continue

        session.delete(row)
        session.commit()
        return {"mensaje": "Promoción eliminada definitivamente."}

    # ==========================================================
    # RACKNOVA POS V5: REPORTE DIARIO DETALLADO
    # ==========================================================

    @app.get("/pos/v5/reportes/diario")
    def daily_report_v5(
        fecha: str,
        caja: Optional[str] = None,
        operador: Optional[str] = None,
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),

        rn_empresa_id: str | None = Header(default=None, alias="X-Empresa-ID"),):
        _rn_bind_empresa(session, current_user, rn_empresa_id, allowed_roles={'admin', 'owner'})  # RACKNOVA_MULTIEMPRESA_FASE2_LOCAL_FIRST
        try:
            report_date = date.fromisoformat(fecha)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail="Fecha inválida. Usa YYYY-MM-DD.",
            ) from exc

        return _v5_daily_report(
            session,
            report_date,
            Movimiento,
            caja,
            operador,
        )

    # ==========================================================
    # RACKNOVA_MULTIEMPRESA_FASE1 - API DE EMPRESAS
    # ==========================================================

    @app.get("/multiempresa/mis-empresas")
    def multiempresa_my_companies_f1(
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),
    ):
        return {
            "empresas": _multi_memberships(session, current_user),
            "empresa_principal": MULTIEMPRESA_DEFAULT_ID,
            "selector_habilitado": False,
        }

    @app.get("/multiempresa/contexto")
    def multiempresa_context_f1(
        empresa_id: Optional[str] = None,
        session: Session = Depends(get_session),
        current_user: Any = Depends(operator_user),
    ):
        return _multi_empresa_context(
            session,
            current_user,
            empresa_id,
        )

    @app.post("/multiempresa/empresas")
    def multiempresa_create_company_f1(
        payload: Dict[str, Any],
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),
    ):
        company = _multi_create_company(session, current_user, payload)
        return {
            "mensaje": "Empresa creada. Todavía no se habilita el cambio de empresa hasta completar la tenantización de módulos.",
            "empresa": company,
        }

    @app.post("/multiempresa/empresas/{empresa_id}/usuarios")
    def multiempresa_add_user_f1(
        empresa_id: str,
        payload: Dict[str, Any],
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),
    ):
        membership = _multi_add_member(
            session,
            current_user,
            empresa_id,
            payload,
        )
        return {
            "mensaje": "Usuario vinculado a la empresa.",
            "membresia": membership,
        }

    @app.get("/multiempresa/diagnostico")
    def multiempresa_diagnostic_f1(
        session: Session = Depends(get_session),
        current_user: Any = Depends(admin_user),
    ):
        return _multi_diagnostic(session)



